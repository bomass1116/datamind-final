"""
FILE    : backend/export_report.py
KEGUNAAN: Export PDF laporan profesional (4-5 halaman), Excel multi-sheet, HTML dashboard
"""

import os
import json
from datetime import datetime

import pandas as pd

from backend.descriptive_stats import get_numerical_stats
from backend.categorical_analysis import get_categorical_stats

OUTPUT_DIR = 'outputs'
os.makedirs(os.path.join(OUTPUT_DIR, 'reports'), exist_ok=True)
os.makedirs(os.path.join(OUTPUT_DIR, 'exported_files'), exist_ok=True)



# Data tim pengembang
TEAM_MEMBERS = [
    {"name": "Den Yuan Frasseka",       "nim": "52250050", "role": "Project Lead & Koordinator Analisis Data"},
    {"name": "Boma Satrio Wicaksono D.","nim": "52250061", "role": "Backend Development & Data Pipeline"},
    {"name": "Frizzy Litmensyah",       "nim": "52250062", "role": "EDA & Pembuatan Visualisasi Grafik"},
    {"name": "Angelica Florentina M",   "nim": "52250063", "role": "UI/UX Design Dashboard & Frontend"},
    {"name": "Adam Richie Wijaya",      "nim": "52250064", "role": "Statistical Analysis & Insight Generator"},
    {"name": "Andre",                   "nim": "52250065", "role": "Data Cleaning, Preprocessing & Validasi"},
]
DOSEN      = "Bakti Siregar, M.Sc., CDS"
INSTITUSI  = "Institut Teknologi Sains Bandung"
PRODI      = "Program Studi Sains Data"
KELOMPOK   = "Kelompok 4 · Kelas B"
TAHUN      = "2026"


def export_pdf(df, col_meta, pages=None, chart_data=None, lang='id'):
    if pages is None:
        pages = ['cover', 'tim', 'daftar_isi', 'bab1', 'bab2', 'bab3', 'bab4']
    if chart_data is None:
        chart_data = {}

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, PageBreak, KeepTogether, KeepInFrame
    )
    from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT, TA_RIGHT

    filepath = os.path.join(OUTPUT_DIR, 'reports', 'DataMind_Laporan.pdf')

    # Nomor halaman
    page_numbers = {}
    _page_counter = [2]  # cover = 1, daftar isi = 2, dst

    doc = SimpleDocTemplate(
        filepath, pagesize=A4,
        leftMargin=3*cm, rightMargin=2.5*cm,
        topMargin=2.5*cm, bottomMargin=2.5*cm
    )

    now        = datetime.now()
    num_stats  = get_numerical_stats(df, col_meta)
    cat_stats  = get_categorical_stats(df, col_meta)

    total_rows    = len(df)
    total_cols    = len(df.columns)
    num_cols      = sum(1 for v in col_meta.values() if v in ['interval','ratio'])
    cat_cols      = sum(1 for v in col_meta.values() if v in ['nominal','ordinal'])
    dt_cols       = sum(1 for v in col_meta.values() if v == 'datetime')
    missing_cells = int(df.isnull().sum().sum())
    missing_pct   = round(missing_cells / max(df.size, 1) * 100, 2)

    # ── Warna
    C_NAVY  = colors.HexColor('#1B2A4A')
    C_BLUE  = colors.HexColor('#6366F1')
    C_BLUE2 = colors.HexColor('#EEF2FF')
    C_GREEN = colors.HexColor('#059669')
    C_GRN2  = colors.HexColor('#D1FAE5')
    C_GRAY  = colors.HexColor('#64748B')
    C_LIGHT = colors.HexColor('#F8FAFC')
    C_LINE  = colors.HexColor('#CBD5E1')
    C_GOLD  = colors.HexColor('#D97706')
    WHITE   = colors.white

    # ── Style helper
    def ps(name, **kw):
        return ParagraphStyle(name, fontName=kw.pop('font','Times-Roman'),
                              fontSize=kw.pop('size',11), leading=kw.pop('leading',16),
                              textColor=kw.pop('color',C_NAVY), **kw)

    S = {
        'cover_title': ps('ct', font='Times-Bold', size=22, color=C_BLUE,
                          alignment=TA_CENTER, spaceAfter=6, leading=28),
        'cover_sub':   ps('cs', size=13, color=C_GRAY, alignment=TA_CENTER, spaceAfter=6),
        'cover_inst':  ps('ci', font='Times-Bold', size=12, color=C_NAVY,
                          alignment=TA_CENTER, spaceAfter=4),
        'cover_date':  ps('cd', size=11, color=C_GRAY, alignment=TA_CENTER),
        'bab_label':   ps('bl', font='Times-Bold', size=10, color=C_BLUE,
                           spaceBefore=6, spaceAfter=3),
        'bab':         ps('bab', font='Times-Bold', size=13, color=C_BLUE,
                           spaceBefore=14, spaceAfter=6),
        'sub':         ps('sub', font='Times-Bold', size=11, color=C_NAVY,
                          spaceBefore=10, spaceAfter=4),
        'body':        ps('body', size=11, leading=18, spaceAfter=6,
                          alignment=TA_JUSTIFY),
        'tbl_hdr':     ps('th', font='Times-Bold', size=9, color=WHITE),
        'caption':     ps('cap', size=9, color=C_GRAY, alignment=TA_CENTER, spaceAfter=8),
        'footer':      ps('ft', size=9, color=C_GRAY, alignment=TA_CENTER),
        'daftar_bab':  ps('db', font='Times-Bold', size=11, color=C_NAVY,
                          spaceBefore=8, spaceAfter=2),
        'daftar_sub':  ps('ds', size=10, color=C_GRAY, spaceBefore=2, spaceAfter=1,
                          leftIndent=16),
        'toc_title':   ps('tt', font='Times-Bold', size=16, color=C_BLUE,
                          alignment=TA_CENTER, spaceAfter=16),
        'tim_name':    ps('tn', font='Times-Bold', size=11, color=C_NAVY,
                          spaceAfter=1),
        'tim_nim':     ps('tnim', size=10, color=C_GRAY, spaceAfter=1),
        'tim_role':    ps('tr', size=10, color=C_BLUE, spaceAfter=4),
    }

    def hr(color=C_LINE, thick=0.5):
        return HRFlowable(width='100%', thickness=thick, color=color,
                          spaceAfter=10, spaceBefore=4)

    def tbl_style(hdr_color=C_NAVY, alt=C_LIGHT, font_size=9):
        return TableStyle([
            ('BACKGROUND',    (0,0),(-1,0), hdr_color),
            ('TEXTCOLOR',     (0,0),(-1,0), WHITE),
            ('FONTNAME',      (0,0),(-1,0), 'Times-Bold'),
            ('FONTSIZE',      (0,0),(-1,-1), font_size),
            ('FONTNAME',      (0,1),(-1,-1), 'Times-Roman'),
            ('GRID',          (0,0),(-1,-1), 0.4, C_LINE),
            ('ROWBACKGROUNDS',(0,1),(-1,-1), [alt, WHITE]),
            ('LEFTPADDING',   (0,0),(-1,-1), 5),
            ('RIGHTPADDING',  (0,0),(-1,-1), 5),
            ('TOPPADDING',    (0,0),(-1,-1), 5),
            ('BOTTOMPADDING', (0,0),(-1,-1), 5),
            ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
            ('WORDWRAP',      (0,0),(-1,-1), True),
        ])

    _T = lambda id, en: en if lang=='en' else id

    def fmt(v, decimals=4):
        if v is None or v == '' or v == '-': return '-'
        try:
            n = float(v)
        except (ValueError, TypeError):
            return str(v)[:12]
        if abs(n) >= 1e9:  return f'{n/1e9:.2f}B'
        if abs(n) >= 1e6:  return f'{n/1e6:.2f}M'
        if abs(n) >= 1e4:  return f'{n:,.0f}'
        return f'{n:.{decimals}g}'

    elems = []

    # ═══════════════════════════════════════════════════════
    # HALAMAN 1 — COVER
    # ═══════════════════════════════════════════════════════
    if 'cover' in pages:
        elems.append(Spacer(1, 1.8*cm))
        elems.append(HRFlowable(width='100%', thickness=4, color=C_BLUE,
                                spaceAfter=18))
        elems.append(Paragraph('EXPLORATORY DATA ANALYSIS REPORT' if lang=='en' else 'LAPORAN ANALISIS DATA EKSPLORASI', S['cover_title']))
        elems.append(Paragraph('Exploratory Data Analysis (EDA)', S['cover_sub']))
        elems.append(Spacer(1, 0.5*cm))
        elems.append(HRFlowable(width='55%', thickness=1, color=C_LINE, spaceAfter=16))
        elems.append(Spacer(1, 0.3*cm))

        cover_rows = [
            [Paragraph('<b>'+_T('Keterangan','Description')+'</b>', S['tbl_hdr']),
             Paragraph('<b>Detail</b>', S['tbl_hdr'])],
            [ _T('Dataset','Dataset'), f'{df.columns[0]} ... ({total_cols} '+_T('kolom','columns')+')'],
            [ _T('Total Observasi','Total Observations'),  f'{total_rows:,} '+_T('baris','rows')],
            [ _T('Total Variabel','Total Variables'),   f'{total_cols} '+_T('kolom','columns')],
            [ _T('Var. Numerik','Num. Variables'),     f'{num_cols} '+_T('kolom','columns')],
            [ _T('Var. Kategorikal','Cat. Variables'), f'{cat_cols} '+_T('kolom','columns')],
            [ _T('Var. Datetime','Datetime Variables'),    f'{dt_cols} '+_T('kolom','columns')],
            ['Missing Values',   f'{missing_cells:,} '+_T('sel','cells')+f' ({missing_pct}%)'],
            [ _T('Tanggal Laporan','Report Date'),  now.strftime(_T('%d %B %Y','%B %d, %Y'))],
            [ _T('Waktu Generate','Generated At'),   now.strftime(_T('%H:%M WIB','%H:%M %Z')).replace('WIB',_T('WIB','GMT+7'))],
        ]
        cover_tbl = Table(cover_rows, colWidths=[6*cm, 9*cm])
        cover_tbl.setStyle(tbl_style(C_NAVY, C_LIGHT))
        elems.append(cover_tbl)
        elems.append(Spacer(1, 1.2*cm))
        elems.append(HRFlowable(width='100%', thickness=4, color=C_BLUE, spaceAfter=12))
        elems.append(Paragraph(INSTITUSI, S['cover_inst']))
        elems.append(Paragraph(f'{PRODI} — {KELOMPOK}', S['cover_inst']))
        elems.append(Paragraph(_T('Dosen:','Lecturer:') + f' {DOSEN}', S['cover_date']))
        elems.append(Paragraph(f'© {TAHUN} DataMind. '+_T('Seluruh hak dilindungi.','All rights reserved.'), S['cover_date']))
        elems.append(PageBreak())

    # ═══════════════════════════════════════════════════════
    # HALAMAN 2 — DAFTAR ISI
    # ═══════════════════════════════════════════════════════
    if 'daftar_isi' in pages:
        elems.append(Paragraph(_T('DAFTAR ISI', 'TABLE OF CONTENTS'), S['toc_title']))
        elems.append(hr(C_BLUE, 2))

        def toc_row(label, page, bold=False):
            fn = 'Times-Bold' if bold else 'Times-Roman'
            sz = 11 if bold else 10
            col_l = C_NAVY if bold else C_GRAY
            return Table([[
                Paragraph(f'<font name="{fn}" size="{sz}" color="#{col_l.hexval()[2:]}">{label}</font>',
                          ps(f'toc_{label[:5]}', font=fn, size=sz, color=col_l)),
                Paragraph(f'<font name="{fn}" size="{sz}">{page}</font>',
                          ps(f'toc_p{page}', font=fn, size=sz, color=col_l,
                             alignment=TA_RIGHT)),
            ]], colWidths=[13.5*cm, 2*cm],
                style=TableStyle([
                    ('LINEBELOW', (0,0), (0,0), 0.3,
                     colors.HexColor('#CBD5E1') if not bold else colors.HexColor('#6366F1')),
                    ('TOPPADDING', (0,0),(-1,-1), 4 if bold else 2),
                    ('BOTTOMPADDING', (0,0),(-1,-1), 4 if bold else 2),
                ]))

        toc_items = [
            ('Cover',                                                         '1',  True),
            (_T('Daftar Isi','Table of Contents'),                            '2',  True),
            (_T('BAB I — Pendahuluan dan Gambaran Umum Dataset','CHAPTER I — Introduction and Dataset Overview'),  '3',  True),
            (_T('  1.1 Latar Belakang','  1.1 Background'),                   '3',  False),
            (_T('  1.2 Tujuan Analisis','  1.2 Analysis Objectives'),         '3',  False),
            (_T('  1.3 Ringkasan Dataset','  1.3 Dataset Summary'),           '3',  False),
            (_T('  1.4 Pratinjau Data (5 Baris Pertama)','  1.4 Data Preview (First 5 Rows)'),  '3',  False),
            (_T('BAB II — Statistik Deskriptif Variabel Numerik','CHAPTER II — Descriptive Statistics Numerical Variables'), '4',  True),
            (_T('  2.1 Ringkasan Statistik Numerik','  2.1 Numerical Statistics Summary'),       '4',  False),
            (_T('  2.2 Analisis Distribusi dan Normalitas','  2.2 Distribution and Normality Analysis'), '4',  False),
            (_T('  2.3 Analisis Outlier','  2.3 Outlier Analysis'),            '4',  False),
            (_T('BAB III — Variabel Kategorikal & Kualitas Data','CHAPTER III — Categorical Variables & Data Quality'), '5',  True),
            (_T('  3.1 Analisis Variabel Kategorikal','  3.1 Categorical Variable Analysis'),     '5',  False),
            (_T('  3.2 Pola Dominasi Kategori','  3.2 Category Dominance Patterns'),            '5',  False),
            (_T('  3.3 Analisis Kualitas Data Komprehensif','  3.3 Comprehensive Data Quality Analysis'),   '5',  False),
            (_T('BAB IV — Temuan Utama & Rekomendasi Strategis','CHAPTER IV — Key Findings & Strategic Recommendations'),  '6',  True),
            (_T('  4.1 Ringkasan Temuan Utama','  4.1 Summary of Key Findings'),            '6',  False),
            (_T('  4.2 Rekomendasi Analisis Lanjutan','  4.2 Recommendations for Further Analysis'),     '6',  False),
            (_T('Halaman Tim Pengembang','Development Team Page'),              '1',  True),
            (_T('Lampiran — Visualisasi Data Pilihan','Appendix — Selected Data Visualizations'), 'Lamp.', True),
        ]
        for label, page, bold in toc_items:
            elems.append(toc_row(label, page, bold))

        elems.append(Spacer(1, 0.8*cm))
        elems.append(Paragraph(
            _T(f'Laporan ini terdiri dari <b>8 halaman utama</b> (termasuk Tim Pengembang di Halaman 1) ditambah lampiran visualisasi. '
               f'Dibuat otomatis oleh <b>DataMind EDA Platform</b> pada {now.strftime("%d %B %Y pukul %H:%M WIB")}.',
               f'This report consists of <b>8 main pages</b> (including Development Team on Page 1) plus visualization appendix. '
               f'Automatically generated by <b>DataMind EDA Platform</b> on {now.strftime("%B %d, %Y at %H:%M WIB")}.'), S['body']))
        elems.append(PageBreak())

    # ═══════════════════════════════════════════════════════
    # HALAMAN 3 — BAB I
    # ═══════════════════════════════════════════════════════
    if 'bab1' in pages:
        elems.append(Paragraph('CHAPTER I' if lang=='en' else 'BAB I', S['bab']))
        elems.append(Paragraph('INTRODUCTION AND DATASET OVERVIEW' if lang=='en' else 'PENDAHULUAN DAN GAMBARAN UMUM DATASET', S['bab']))
        elems.append(hr(C_BLUE, 1.5))

        elems.append(Paragraph(_T('1.1 Latar Belakang','1.1 Background'), S['sub']))
        elems.append(Paragraph(
            _T(f'Laporan ini disusun sebagai hasil analisis eksplorasi data (<i>Exploratory Data Analysis</i>/EDA) '
               f'menggunakan platform DataMind — Intelligent EDA Platform yang dikembangkan oleh Kelompok 4 '
               f'Program Studi Sains Data, {INSTITUSI}. '
               f'Analisis dilakukan terhadap dataset yang memiliki <b>{total_rows:,} observasi</b> dan '
               f'<b>{total_cols} variabel</b>, terdiri dari {num_cols} variabel numerik, '
               f'{cat_cols} variabel kategorikal, dan {dt_cols} variabel datetime. '
               f'EDA merupakan tahapan krusial dalam siklus ilmu data yang bertujuan memahami '
               f'struktur, distribusi, kualitas, dan pola tersembunyi di dalam data sebelum '
               f'dilakukan pemodelan statistik maupun machine learning.',
               f'This report is prepared as the result of exploratory data analysis (EDA) '
               f'using the DataMind — Intelligent EDA Platform developed by Group 4 '
               f'of the Data Science Study Program, {INSTITUSI}. '
               f'The analysis was conducted on a dataset with <b>{total_rows:,} observations</b> and '
               f'<b>{total_cols} variables</b>, consisting of {num_cols} numeric variables, '
               f'{cat_cols} categorical variables, and {dt_cols} datetime variables. '
               f'EDA is a crucial stage in the data science lifecycle aimed at understanding '
               f'the structure, distribution, quality, and hidden patterns within the data before '
               f'proceeding to statistical modeling or machine learning.'), S['body']))

        elems.append(Paragraph(_T('1.2 Tujuan Analisis','1.2 Analysis Objectives'), S['sub']))
        elems.append(Paragraph(
            _T(f'Tujuan utama laporan ini adalah: (1) mengidentifikasi karakteristik statistik deskriptif '
               f'setiap variabel, (2) mendeteksi dan mendokumentasikan masalah kualitas data seperti '
               f'missing values dan outlier, (3) menganalisis distribusi dan normalitas data numerik, '
               f'(4) memahami pola dan frekuensi pada variabel kategorikal, serta '
               f'(5) memberikan rekomendasi strategis untuk analisis lanjutan yang dapat ditindaklanjuti '
               f'oleh tim riset maupun pemangku kepentingan.',
               f'The main objectives of this report are: (1) identify descriptive statistical characteristics '
               f'of each variable, (2) detect and document data quality issues such as '
               f'missing values and outliers, (3) analyze the distribution and normality of numeric data, '
               f'(4) understand patterns and frequencies in categorical variables, and '
               f'(5) provide strategic recommendations for further analysis that can be followed up '
               f'by the research team and stakeholders.'), S['body']))

        elems.append(Paragraph(_T('1.3 Ringkasan Dataset','1.3 Dataset Summary'), S['sub']))
        elems.append(Paragraph(
            _T(f'Dataset yang dianalisis memiliki dimensi <b>{total_rows:,} baris × {total_cols} kolom</b>. '
               f'Distribusi tipe variabel: {num_cols} variabel numerik (skala interval/rasio), '
               f'{cat_cols} variabel kategorikal (skala nominal/ordinal), '
               f'dan {dt_cols} variabel bertipe datetime. '
               f'Total missing values yang terdeteksi adalah <b>{missing_cells:,} sel</b> atau setara '
               f'<b>{missing_pct}%</b> dari keseluruhan data — tergolong '
               f'<b>{"rendah dan dapat diabaikan" if missing_pct < 5 else "cukup signifikan dan memerlukan penanganan"}</b>.',
               f'The analyzed dataset has dimensions of <b>{total_rows:,} rows × {total_cols} columns</b>. '
               f'Variable type distribution: {num_cols} numeric variables (interval/ratio scale), '
               f'{cat_cols} categorical variables (nominal/ordinal scale), '
               f'and {dt_cols} datetime variables. '
               f'Total missing values detected are <b>{missing_cells:,} cells</b> or equivalent to '
               f'<b>{missing_pct}%</b> of the entire data — classified as '
               f'<b>{"low and negligible" if missing_pct < 5 else "quite significant and requires handling"}</b>.'), S['body']))

        # Tabel ringkasan dataset
        sum_rows = [
            [Paragraph('<b>'+_T('Metrik','Metric')+'</b>', S['tbl_hdr']),
             Paragraph('<b>'+_T('Nilai','Value')+'</b>', S['tbl_hdr']),
             Paragraph('<b>'+_T('Keterangan','Description')+'</b>', S['tbl_hdr'])],
            [_T('Total Baris','Total Rows'),       f'{total_rows:,}',         _T('Jumlah observasi/record','Number of observations/records')],
            [_T('Total Kolom','Total Columns'),       f'{total_cols}',           _T('Jumlah fitur/variabel','Number of features/variables')],
            [_T('Variabel Numerik','Numeric Variables'),  f'{num_cols}',             _T('Skala interval & rasio','Interval & ratio scale')],
            [_T('Variabel Kategorikal','Categorical Variables'), f'{cat_cols}',          _T('Skala nominal & ordinal','Nominal & ordinal scale')],
            [_T('Variabel Datetime','Datetime Variables'), f'{dt_cols}',              _T('Tipe tanggal/waktu','Date/time type')],
            [_T('Total Missing','Total Missing'),     f'{missing_cells:,} '+_T('sel','cells'),  f'{missing_pct}% '+_T('dari total data','of total data')],
        ]
        sum_tbl = Table(sum_rows, colWidths=[5*cm, 4*cm, 6.5*cm], repeatRows=1)
        sum_tbl.setStyle(tbl_style(C_NAVY, C_LIGHT))
        elems.append(sum_tbl)
        elems.append(Paragraph(_T('Tabel 1.1 Ringkasan dimensi dan tipe variabel dataset','Table 1.1 Summary of dataset dimensions and variable types'), S['caption']))

        elems.append(Paragraph(_T('1.4 Pratinjau Data (5 Baris Pertama)','1.4 Data Preview (First 5 Rows)'), S['sub']))
        preview   = df.head(5).where(df.head(5).notna(), other='-')
        max_cols  = min(7, len(df.columns))
        prev_cols = list(df.columns[:max_cols])
        col_w     = [15.5 / max_cols * cm] * max_cols
        prev_data = [[Paragraph(f'<b>{str(c)[:11]}</b>', S['tbl_hdr']) for c in prev_cols]]
        for i in range(len(preview)):
            prev_data.append([
                Paragraph(str(preview.iloc[i][c])[:15], ps(f'pr{i}{j}', size=8))
                for j, c in enumerate(prev_cols)
            ])
        prev_tbl = Table(prev_data, colWidths=col_w, repeatRows=1)
        prev_tbl.setStyle(tbl_style(C_NAVY, C_LIGHT, font_size=8))
        elems.append(prev_tbl)
        elems.append(Paragraph(_T('Tabel 1.2 Pratinjau lima baris pertama dataset','Table 1.2 Preview of the first five rows of the dataset'), S['caption']))
        elems.append(PageBreak())

    # ═══════════════════════════════════════════════════════
    # HALAMAN 4 — BAB II: STATISTIK NUMERIK
    # ═══════════════════════════════════════════════════════
    if 'bab2' in pages:
        elems.append(Paragraph('CHAPTER II' if lang=='en' else 'BAB II', S['bab']))
        elems.append(Paragraph('DESCRIPTIVE STATISTICS — NUMERICAL VARIABLES' if lang=='en' else 'STATISTIK DESKRIPTIF VARIABEL NUMERIK', S['bab']))
        elems.append(hr(C_BLUE, 1.5))

        elems.append(Paragraph(_T('2.1 Ringkasan Statistik Numerik','2.1 Numerical Statistics Summary'), S['sub']))
        elems.append(Paragraph(
            _T(f'Tabel berikut menyajikan statistik deskriptif lengkap untuk <b>{num_cols} variabel numerik</b>. '
               f'Metrik meliputi ukuran pemusatan (mean, median), penyebaran (standar deviasi, min, max), '
               f'bentuk distribusi (skewness, kurtosis), serta hasil uji normalitas. '
               f'Nilai yang sangat besar diformat ringkas (M = juta, B = miliar) agar tabel tidak terpotong.',
               f'The following table presents complete descriptive statistics for <b>{num_cols} numeric variables</b>. '
               f'Metrics include central tendency (mean, median), dispersion (standard deviation, min, max), '
               f'distribution shape (skewness, kurtosis), and normality test results. '
               f'Very large values are formatted concisely (M = million, B = billion) to prevent table truncation.'),
            S['body']))

        if num_stats:
            num_hdr = [
                Paragraph('<b>'+_T('Variabel','Variable')+'</b>', S['tbl_hdr']),
                Paragraph('<b>'+_T('Skala','Scale')+'</b>',   S['tbl_hdr']),
                Paragraph('<b>N</b>',       S['tbl_hdr']),
                Paragraph('<b>Mean</b>',    S['tbl_hdr']),
                Paragraph('<b>Median</b>',  S['tbl_hdr']),
                Paragraph('<b>Std</b>',     S['tbl_hdr']),
                Paragraph('<b>Min</b>',     S['tbl_hdr']),
                Paragraph('<b>Max</b>',     S['tbl_hdr']),
                Paragraph('<b>Skew</b>',    S['tbl_hdr']),
                Paragraph('<b>'+_T('Outlier','Outlier')+'</b>', S['tbl_hdr']),
                Paragraph('<b>'+_T('Normal','Normal')+'</b>',  S['tbl_hdr']),
            ]
            num_data = [num_hdr]
            cell_s = ps('nc', size=8, leading=10)
            for r in num_stats:
                norm = str(r.get('normality', '-'))
                norm_short = _T('Ya','Yes') if norm == 'Normal' else _T('Tidak','No')
                num_data.append([
                    Paragraph(str(r.get('column','-'))[:16], cell_s),
                    Paragraph(str(r.get('scale','-'))[:6].upper(), cell_s),
                    Paragraph(str(r.get('count','-')), cell_s),
                    Paragraph(fmt(r.get('mean')), cell_s),
                    Paragraph(fmt(r.get('median')), cell_s),
                    Paragraph(fmt(r.get('std')), cell_s),
                    Paragraph(fmt(r.get('min')), cell_s),
                    Paragraph(fmt(r.get('max')), cell_s),
                    Paragraph(fmt(r.get('skewness')), cell_s),
                    Paragraph(str(r.get('outliers','0')), cell_s),
                    Paragraph(norm_short, cell_s),
                ])
            cw = [2.8*cm, 1.2*cm, 0.9*cm, 1.6*cm, 1.6*cm,
                  1.6*cm, 1.6*cm, 1.6*cm, 1.2*cm, 1.2*cm, 1.2*cm]
            num_tbl = Table(num_data, colWidths=cw, repeatRows=1)
            ts = tbl_style(C_BLUE, C_BLUE2, font_size=8)
            ts.add('ALIGN', (2,0), (-1,-1), 'CENTER')
            num_tbl.setStyle(ts)
            elems.append(num_tbl)
            elems.append(Paragraph(_T('Tabel 2.1 Statistik deskriptif variabel numerik (tabel penuh, tidak terpotong)',
                                      'Table 2.1 Descriptive statistics of numeric variables (full table, not truncated)'), S['caption']))

        not_normal   = [r for r in num_stats if str(r.get('normality','')).lower().replace('_',' ') in ['not normal']]
        normal_ok    = [r for r in num_stats if r.get('normality') == 'Normal']
        skewed_pos   = [r for r in num_stats if float(r.get('skewness') or 0) >  1]
        skewed_neg   = [r for r in num_stats if float(r.get('skewness') or 0) < -1]
        outlier_cols = [r for r in num_stats if float(r.get('outliers') or 0) > 0]

        elems.append(Paragraph(_T('2.2 Analisis Distribusi dan Normalitas','2.2 Distribution and Normality Analysis'), S['sub']))
        dt = _T(f'Hasil uji normalitas menunjukkan <b>{len(normal_ok)} dari {num_cols} variabel numerik</b> '
                f'berdistribusi normal, sementara <b>{len(not_normal)} variabel</b> tidak normal. ',
                f'Normality test results show <b>{len(normal_ok)} out of {num_cols} numeric variables</b> '
                f'are normally distributed, while <b>{len(not_normal)} variables</b> are not normal. ')
        if not_normal:
            dt += _T(f'Variabel tidak normal: <b>{", ".join([r["column"] for r in not_normal[:5]])}</b>. '
                     f'Kondisi ini mengindikasikan perlunya transformasi data (log, sqrt) atau '
                     f'penggunaan metode statistik non-parametrik. ',
                     f'Non-normal variables: <b>{", ".join([r["column"] for r in not_normal[:5]])}</b>. '
                     f'This indicates the need for data transformation (log, sqrt) or '
                     f'the use of non-parametric statistical methods. ')
        if skewed_pos:
            dt += _T(f'Terdapat {len(skewed_pos)} variabel <i>right-skewed</i> '
                     f'(skewness > 1): {", ".join([r["column"] for r in skewed_pos[:3]])}. ',
                     f'There are {len(skewed_pos)} <i>right-skewed</i> variables '
                     f'(skewness > 1): {", ".join([r["column"] for r in skewed_pos[:3]])}. ')
        if skewed_neg:
            dt += _T(f'Sementara {len(skewed_neg)} variabel <i>left-skewed</i> '
                     f'(skewness < −1): {", ".join([r["column"] for r in skewed_neg[:3]])}. ',
                     f'Meanwhile {len(skewed_neg)} <i>left-skewed</i> variables '
                     f'(skewness < −1): {", ".join([r["column"] for r in skewed_neg[:3]])}. ')
        elems.append(Paragraph(dt, S['body']))

        elems.append(Paragraph(_T('2.3 Analisis Outlier','2.3 Outlier Analysis'), S['sub']))
        if outlier_cols:
            out_rows = [
                [Paragraph('<b>'+_T('Variabel','Variable')+'</b>', S['tbl_hdr']),
                 Paragraph('<b>'+_T('Jumlah Outlier','Outlier Count')+'</b>', S['tbl_hdr']),
                 Paragraph('<b>'+_T('% dari Total','% of Total')+'</b>', S['tbl_hdr']),
                 Paragraph('<b>'+_T('Rekomendasi','Recommendation')+'</b>', S['tbl_hdr'])],
            ]
            for r in outlier_cols:
                cnt  = int(r.get('outliers', 0))
                pct_ = round(cnt / max(total_rows, 1) * 100, 2)
                rec  = _T('Winsorize / Transformasi Log','Winsorize / Log Transform') if pct_ > 5 else _T('Investigasi Lebih Lanjut','Further Investigation')
                out_rows.append([
                    Paragraph(str(r['column']), ps('oc', size=9)),
                    Paragraph(str(cnt), ps('ov', size=9)),
                    Paragraph(f'{pct_}%', ps('op', size=9)),
                    Paragraph(rec, ps('or', size=9)),
                ])
            out_tbl = Table(out_rows, colWidths=[4*cm, 3.5*cm, 3*cm, 5*cm], repeatRows=1)
            out_tbl.setStyle(tbl_style(C_GOLD, C_LIGHT))
            elems.append(out_tbl)
            elems.append(Paragraph(_T('Tabel 2.2 Ringkasan outlier per variabel numerik','Table 2.2 Summary of outliers per numeric variable'), S['caption']))
        else:
            elems.append(Paragraph(
                _T('Analisis outlier menggunakan metode IQR tidak mendeteksi outlier yang signifikan. '
                   'Dataset dalam kondisi bersih dari nilai ekstrem yang mengganggu analisis.',
                   'Outlier analysis using the IQR method detected no significant outliers. '
                   'The dataset is clean from extreme values that could disrupt the analysis.'), S['body']))
        elems.append(PageBreak())

    # ═══════════════════════════════════════════════════════
    # HALAMAN 5 — BAB III: KATEGORIKAL & KUALITAS DATA
    # ═══════════════════════════════════════════════════════
    if 'bab3' in pages:
        elems.append(Paragraph('CHAPTER III' if lang=='en' else 'BAB III', S['bab']))
        elems.append(Paragraph('CATEGORICAL VARIABLES & DATA QUALITY' if lang=='en' else 'STATISTIK DESKRIPTIF VARIABEL KATEGORIKAL DAN KUALITAS DATA', S['bab']))
        elems.append(hr(C_BLUE, 1.5))

        elems.append(Paragraph(_T('3.1 Analisis Variabel Kategorikal','3.1 Categorical Variable Analysis'), S['sub']))
        elems.append(Paragraph(
            _T(f'Dataset memiliki <b>{cat_cols} variabel kategorikal</b> yang dianalisis dari aspek '
               f'frekuensi, distribusi kategori, dan dominasi nilai modus. '
               f'Pemahaman terhadap variabel kategorikal penting untuk analisis segmentasi, '
               f'pemodelan klasifikasi, dan interpretasi bisnis yang tepat.',
               f'The dataset has <b>{cat_cols} categorical variables</b> analyzed in terms of '
               f'frequency, category distribution, and mode dominance. '
               f'Understanding categorical variables is important for segmentation analysis, '
               f'classification modeling, and proper business interpretation.'), S['body']))

        if cat_stats:
            cat_hdr = [
                Paragraph('<b>'+_T('Variabel','Variable')+'</b>',  S['tbl_hdr']),
                Paragraph('<b>'+_T('Skala','Scale')+'</b>',     S['tbl_hdr']),
                Paragraph('<b>N</b>',         S['tbl_hdr']),
                Paragraph('<b>'+_T('Unik','Unique')+'</b>',      S['tbl_hdr']),
                Paragraph('<b>'+_T('Modus','Mode')+'</b>',     S['tbl_hdr']),
                Paragraph('<b>'+_T('Frek.','Freq.')+'</b>',     S['tbl_hdr']),
                Paragraph('<b>'+_T('Modus%','Mode%')+'</b>',    S['tbl_hdr']),
                Paragraph('<b>'+_T('Missing%','Missing%')+'</b>',  S['tbl_hdr']),
            ]
            cat_data = [cat_hdr]
            cell_s   = ps('cc', size=9, leading=11)
            for r in cat_stats:
                cat_data.append([
                    Paragraph(str(r.get('column','-'))[:15], cell_s),
                    Paragraph(str(r.get('scale','-'))[:6].upper(), cell_s),
                    Paragraph(str(r.get('count','-')), cell_s),
                    Paragraph(str(r.get('unique','-')), cell_s),
                    Paragraph(str(r.get('mode','-'))[:12], cell_s),
                    Paragraph(str(r.get('mode_freq','-')), cell_s),
                    Paragraph(f"{r.get('mode_pct','-')}%", cell_s),
                    Paragraph(f"{r.get('missing_pct',0)}%", cell_s),
                ])
            cw = [3.2*cm, 1.6*cm, 1.2*cm, 1.4*cm, 3.0*cm, 1.5*cm, 2.0*cm, 1.6*cm]
            cat_tbl = Table(cat_data, colWidths=cw, repeatRows=1)
            ts = tbl_style(C_GREEN, C_GRN2)
            ts.add('ALIGN', (2,0), (-1,-1), 'CENTER')
            cat_tbl.setStyle(ts)
            elems.append(cat_tbl)
            elems.append(Paragraph(_T('Tabel 3.1 Statistik deskriptif variabel kategorikal','Table 3.1 Descriptive statistics of categorical variables'), S['caption']))

        elems.append(Paragraph(_T('3.2 Pola Dominasi Kategori','3.2 Category Dominance Patterns'), S['sub']))
        dominant = [r for r in cat_stats
                    if str(r.get('mode_pct') or '0') not in ('-','','nan')
                    and float(str(r.get('mode_pct') or 0)) > 30]
        if dominant:
            dom = _T(f'Terdapat <b>{len(dominant)} variabel</b> dengan dominasi kategori >30%: ',
                     f'There are <b>{len(dominant)} variables</b> with category dominance >30%: ')
            for r in dominant[:5]:
                dom += _T(f'<b>{r["column"]}</b> didominasi nilai "<i>{r.get("mode","")}</i>" '
                         f'sebesar {r.get("mode_pct","")}%; ',
                         f'<b>{r["column"]}</b> is dominated by "<i>{r.get("mode","")}</i>" '
                         f'at {r.get("mode_pct","")}%; ')
            dom += _T('Dominasi tinggi berpotensi menyebabkan ketidakseimbangan kelas '
                      'dalam pemodelan prediktif dan perlu dipertimbangkan dalam strategi sampling.',
                      'High dominance may cause class imbalance '
                      'in predictive modeling and should be considered in sampling strategy.')
            elems.append(Paragraph(dom, S['body']))
        else:
            elems.append(Paragraph(
                _T('Tidak ditemukan dominasi kategori ekstrem (>30%). '
                   'Distribusi antar kategori relatif seimbang.',
                   'No extreme category dominance (>30%) found. '
                   'Distribution across categories is relatively balanced.'), S['body']))

        elems.append(Paragraph(_T('3.3 Analisis Kualitas Data Komprehensif','3.3 Comprehensive Data Quality Analysis'), S['sub']))
        all_comb    = num_stats + cat_stats
        high_miss   = [r for r in all_comb if float(r.get('missing_pct') or 0) > 10]
        low_miss    = [r for r in all_comb if 0 < float(r.get('missing_pct') or 0) <= 10]
        quality = _T(f'Evaluasi kualitas data dilakukan terhadap seluruh <b>{total_cols} variabel</b>. '
                     f'Total missing values: <b>{missing_cells:,} sel ({missing_pct}%)</b>. ',
                     f'Data quality evaluation was performed on all <b>{total_cols} variables</b>. '
                     f'Total missing values: <b>{missing_cells:,} cells ({missing_pct}%)</b>. ')
        if high_miss:
            quality += _T(f'Variabel dengan missing >10%: '
                          f'<b>{", ".join([r["column"] for r in high_miss[:5]])}</b>. '
                          f'Rekomendasi: imputasi median (numerik) / modus (kategorikal) '
                          f'atau drop jika >30%. ',
                          f'Variables with missing >10%: '
                          f'<b>{", ".join([r["column"] for r in high_miss[:5]])}</b>. '
                          f'Recommendation: median imputation (numeric) / mode imputation (categorical) '
                          f'or drop if >30%. ')
        if low_miss:
            quality += _T(f'Variabel dengan missing rendah (1–10%): '
                          f'{", ".join([r["column"] for r in low_miss[:5]])}. '
                          f'Dapat ditangani dengan imputasi sederhana. ',
                          f'Variables with low missing (1–10%): '
                          f'{", ".join([r["column"] for r in low_miss[:5]])}. '
                          f'Can be handled with simple imputation. ')
        if not high_miss and not low_miss:
            quality += _T('Dataset bersih, tidak ada missing values. ','Dataset is clean, no missing values. ')
        elems.append(Paragraph(quality, S['body']))
        elems.append(PageBreak())

    # ═══════════════════════════════════════════════════════
    # HALAMAN 6 — BAB IV: TEMUAN & REKOMENDASI
    # ═══════════════════════════════════════════════════════
    if 'bab4' in pages:
        elems.append(Paragraph('CHAPTER IV' if lang=='en' else 'BAB IV', S['bab']))
        elems.append(Paragraph('KEY FINDINGS & STRATEGIC RECOMMENDATIONS' if lang=='en' else 'TEMUAN UTAMA DAN REKOMENDASI STRATEGIS', S['bab']))
        elems.append(hr(C_BLUE, 1.5))

        _num  = get_numerical_stats(df, col_meta)
        _cat  = get_categorical_stats(df, col_meta)
        _not_n   = [r for r in _num if str(r.get('normality','')).lower().replace('_',' ') == 'not normal']
        _outlier = [r for r in _num if float(r.get('outliers') or 0) > 0]
        _dom     = [r for r in _cat if str(r.get('mode_pct') or '0') not in ('-','','nan')
                    and float(str(r.get('mode_pct') or 0)) > 30]

        elems.append(Paragraph(_T('4.1 Ringkasan Temuan Utama','4.1 Summary of Key Findings'), S['sub']))
        elems.append(Paragraph(
            _T('Berikut adalah temuan utama yang diidentifikasi dari proses analisis eksplorasi data '
               'yang mencakup aspek statistik deskriptif, kualitas data, distribusi, dan pola kategorikal.',
               'The following are the key findings identified from the exploratory data analysis process '
               'covering descriptive statistics, data quality, distribution, and categorical patterns.'),
            S['body']))

        findings = []
        if _num:
            hm = max(_num, key=lambda r: float(r.get('mean') or 0))
            findings.append(_T(f'Variabel <b>{hm["column"]}</b> memiliki nilai rata-rata tertinggi '
                              f'({fmt(hm.get("mean"))}), menjadikannya variabel dengan magnitudo terbesar.',
                              f'Variable <b>{hm["column"]}</b> has the highest mean value '
                              f'({fmt(hm.get("mean"))}), making it the variable with the largest magnitude.'))
        if _outlier:
            mo = max(_outlier, key=lambda r: float(r.get('outliers') or 0))
            pct_ = round(float(mo.get('outliers',0))/max(total_rows,1)*100, 1)
            findings.append(_T(f'Variabel <b>{mo["column"]}</b> memiliki outlier terbanyak '
                              f'({mo.get("outliers","")} data point / {pct_}%), memerlukan investigasi mendalam.',
                              f'Variable <b>{mo["column"]}</b> has the most outliers '
                              f'({mo.get("outliers","")} data points / {pct_}%), requiring in-depth investigation.'))
        if _not_n:
            findings.append(_T(f'Sebanyak <b>{len(_not_n)} variabel</b> tidak berdistribusi normal: '
                              f'{", ".join([r["column"] for r in _not_n[:4]])}. '
                              f'Disarankan menggunakan korelasi Spearman.',
                              f'A total of <b>{len(_not_n)} variables</b> are not normally distributed: '
                              f'{", ".join([r["column"] for r in _not_n[:4]])}. '
                              f'Spearman correlation is recommended.'))
        if _dom:
            findings.append(_T(f'Variabel <b>{_dom[0]["column"]}</b> menunjukkan dominasi kategori '
                              f'"{_dom[0].get("mode","")}" sebesar {_dom[0].get("mode_pct","")}% '
                              f'— berpotensi menyebabkan class imbalance.',
                              f'Variable <b>{_dom[0]["column"]}</b> shows category dominance '
                              f'of "{_dom[0].get("mode","")}" at {_dom[0].get("mode_pct","")}% '
                              f'— potentially causing class imbalance.'))
        findings.append(_T(f'Rasio missing values <b>{missing_pct}%</b>, tergolong '
                          f'<b>{"sangat baik (<5%)" if missing_pct < 5 else "perlu penanganan sebelum pemodelan"}</b>.',
                          f'Missing values ratio <b>{missing_pct}%</b>, classified as '
                          f'<b>{"very good (<5%)" if missing_pct < 5 else "needs handling before modeling"}</b>.'))

        for i, f in enumerate(findings, 1):
            elems.append(Paragraph(f'{i}. {f}', S['body']))
        elems.append(Spacer(1, 0.3*cm))

        elems.append(Paragraph(_T('4.2 Rekomendasi Analisis Lanjutan','4.2 Recommendations for Further Analysis'), S['sub']))
        recs = [
            (_T('4.2.1 Pra-pemrosesan Data','4.2.1 Data Preprocessing'),
             _T('Tangani outlier dengan winsorization atau transformasi log untuk variabel skewed. '
                'Imputasi missing values: median untuk numerik, modus untuk kategorikal. '
                'Drop kolom/baris jika missing >30%.',
                'Handle outliers with winsorization or log transformation for skewed variables. '
                'Impute missing values: median for numeric, mode for categorical. '
                'Drop columns/rows if missing >30%.')),
            (_T('4.2.2 Analisis Korelasi','4.2.2 Correlation Analysis'),
             _T('Gunakan korelasi Pearson untuk variabel normal dan Spearman untuk non-normal. '
                'Visualisasikan heatmap korelasi untuk mendeteksi multikolinearitas sebelum pemodelan.',
                'Use Pearson correlation for normal variables and Spearman for non-normal. '
                'Visualize correlation heatmap to detect multicollinearity before modeling.')),
            (_T('4.2.3 Segmentasi dan Pengelompokan','4.2.3 Segmentation and Clustering'),
             _T('Manfaatkan variabel kategorikal untuk analisis segmentasi dengan metode K-Means/clustering. '
                'Bandingkan distribusi numerik antar segmen untuk menemukan pola bermakna.',
                'Leverage categorical variables for segmentation analysis using K-Means/clustering. '
                'Compare numeric distributions across segments to find meaningful patterns.')),
            (_T('4.2.4 Pemodelan Prediktif','4.2.4 Predictive Modeling'),
             _T('Terapkan normalisasi/standarisasi pada variabel numerik sebelum pemodelan. '
                'Untuk klasifikasi dengan class imbalance, gunakan SMOTE atau class_weight balancing.',
                'Apply normalization/standardization to numeric variables before modeling. '
                'For classification with class imbalance, use SMOTE or class_weight balancing.')),
            (_T('4.2.5 Validasi dan Reproduktibilitas','4.2.5 Validation and Reproducibility'),
             _T('Dokumentasikan semua keputusan pra-pemrosesan dalam pipeline yang dapat direproduksi. '
                'Lakukan cross-validation (k-fold) untuk memastikan generalisasi model yang baik.',
                'Document all preprocessing decisions in a reproducible pipeline. '
                'Perform cross-validation (k-fold) to ensure good model generalization.')),
        ]
        for title, text in recs:
            elems.append(KeepTogether([
                Paragraph(title, S['sub']),
                Paragraph(text, S['body']),
            ]))

        elems.append(Spacer(1, 0.5*cm))
        elems.append(hr(C_BLUE))
        elems.append(Paragraph(
            _T(f'Laporan dibuat otomatis oleh <b>DataMind Intelligent EDA Platform</b> pada '
               f'{now.strftime("%d %B %Y pukul %H:%M WIB")}. '
               f'{INSTITUSI} — {PRODI} — {KELOMPOK}.',
               f'Report automatically generated by <b>DataMind Intelligent EDA Platform</b> on '
               f'{now.strftime("%B %d, %Y at %H:%M WIB")}. '
               f'{INSTITUSI} — {PRODI} — {KELOMPOK}.'), S['footer']))
        elems.append(PageBreak())

    # ═══════════════════════════════════════════════════════
    # HALAMAN TIM PENGEMBANG
    # ═══════════════════════════════════════════════════════
    if 'tim' in pages:
        elems.append(Paragraph(_T('Halaman 1 — Tim Pengembang','Page 1 — Development Team'), S['bab_label']))
        elems.append(Paragraph(_T('Tim Pengembang','Development Team'), S['bab']))
        elems.append(hr(C_BLUE, 2))

        elems.append(Paragraph(
            _T(f'DataMind dikembangkan oleh mahasiswa {PRODI}, {INSTITUSI} '
               f'sebagai proyek akhir mata kuliah Data Science Programming tahun {TAHUN}.',
               f'DataMind was developed by students of {PRODI}, {INSTITUSI} '
               f'as the final project for the Data Science Programming course, {TAHUN}.'), S['body']))
        elems.append(Spacer(1, 0.3*cm))

        inst_rows = [
            [Paragraph('<b>'+_T('Institusi','Institution')+'</b>', S['tbl_hdr']),
             Paragraph('<b>'+_T('Detail','Detail')+'</b>', S['tbl_hdr'])],
            [_T('Perguruan Tinggi','University'), INSTITUSI],
            [_T('Program Studi','Study Program'),   PRODI],
            [_T('Mata Kuliah','Course'),     'Data Science Programming'],
            [_T('Kelompok','Group'),        KELOMPOK],
            [_T('Dosen Pembimbing','Supervisor'),DOSEN],
            [_T('Tahun Akademik','Academic Year'),  TAHUN],
        ]
        inst_tbl = Table(inst_rows, colWidths=[5*cm, 10.5*cm])
        inst_tbl.setStyle(tbl_style(C_NAVY, C_LIGHT))
        elems.append(inst_tbl)
        elems.append(Paragraph(_T('Tabel Tim.1 Informasi institusi dan mata kuliah','Table Team.1 Institution and course information'), S['caption']))
        elems.append(Spacer(1, 0.5*cm))

        elems.append(Paragraph(_T('Anggota Tim','Team Members'), S['sub']))
        tim_hdr = [
            Paragraph('<b>No.</b>',          S['tbl_hdr']),
            Paragraph('<b>'+_T('Nama Lengkap','Full Name')+'</b>', S['tbl_hdr']),
            Paragraph('<b>NIM</b>',          S['tbl_hdr']),
            Paragraph('<b>'+_T('Kontribusi Utama','Main Contribution')+'</b>', S['tbl_hdr']),
        ]
        tim_data = [tim_hdr]
        cell_s = ps('tmc', size=10, leading=13)
        for i, m in enumerate(TEAM_MEMBERS, 1):
            tim_data.append([
                Paragraph(str(i), cell_s),
                Paragraph(m['name'], cell_s),
                Paragraph(m['nim'], cell_s),
                Paragraph(m['role'], cell_s),
            ])
        tim_tbl = Table(tim_data, colWidths=[0.8*cm, 5.5*cm, 2.8*cm, 6.4*cm], repeatRows=1)
        ts = tbl_style(C_BLUE, C_BLUE2, font_size=10)
        ts.add('ALIGN', (0,0), (0,-1), 'CENTER')
        tim_tbl.setStyle(ts)
        elems.append(tim_tbl)
        elems.append(Paragraph(_T('Tabel Tim.2 Anggota tim dan kontribusi masing-masing','Table Team.2 Team members and their contributions'), S['caption']))
        elems.append(Spacer(1, 0.5*cm))

        elems.append(Paragraph(_T('Tentang DataMind','About DataMind'), S['sub']))
        elems.append(Paragraph(
            _T('DataMind adalah platform analisis data eksplorasi (<i>Exploratory Data Analysis</i>) '
               'berbasis web yang dikembangkan sebagai proyek akhir mata kuliah Data Science Programming. '
               'Platform ini mendukung upload dataset (CSV, XLSX, TXT), analisis statistik otomatis, '
               '20+ jenis visualisasi interaktif, deteksi dan penanganan data quality (missing, duplicate, outlier), '
               'time series analytics, AI-powered insight generation, serta ekspor laporan ke format PDF, Excel, dan HTML. '
               'Seluruh fitur dirancang untuk memudahkan proses EDA bagi data scientist maupun pengguna awam.',
               'DataMind is a web-based exploratory data analysis (EDA) platform '
               'developed as the final project for the Data Science Programming course. '
               'The platform supports dataset upload (CSV, XLSX, TXT), automatic statistical analysis, '
               '20+ types of interactive visualizations, data quality detection and handling (missing, duplicate, outlier), '
               'time series analytics, AI-powered insight generation, and report export to PDF, Excel, and HTML formats. '
               'All features are designed to simplify the EDA process for both data scientists and general users.'),
            S['body']))

    # ═══════════════════════════════════════════════════════
    # LAMPIRAN — VISUALISASI
    # ═══════════════════════════════════════════════════════
    if chart_data:
        import io
        from reportlab.platypus import Image as RLImage

        DL_CHART_LABELS = {
            'histogram':'Histogram','boxplot':'Boxplot','density':'Density Plot',
            'qq_plot':'QQ Plot','violin':'Violin Plot','bar_chart':'Bar Chart',
            'pie_chart':'Pie Chart','count_plot':'Count Plot','pareto_chart':'Pareto Chart',
            'scatter':'Scatter Plot','corr_heatmap':'Correlation Heatmap',
            'regression':'Regression Plot','bubble':'Bubble Chart',
            'boxplot_cat':'Boxplot per Kategori','violin_cat':'Violin per Kategori',
            'grouped_bar':'Grouped Bar','strip_plot':'Strip Plot',
            'ts_line':'Time Series Line','ts_ma':'Moving Average',
            'ts_rolling':'Rolling Mean','ts_trend':'Trend Line',
        }

        elems.append(PageBreak())
        elems.append(Paragraph(_T('LAMPIRAN','APPENDIX'), S['bab']))
        elems.append(Paragraph(_T('VISUALISASI DATA PILIHAN','SELECTED DATA VISUALIZATIONS'), S['bab']))
        elems.append(hr(C_BLUE, 1.5))
        elems.append(Paragraph(
            _T(f'Lampiran ini memuat {len(chart_data)} visualisasi yang dipilih pengguna '
               f'sebagai pelengkap laporan analisis data eksplorasi.',
               f'This appendix contains {len(chart_data)} visualizations selected by the user '
               f'as a supplement to the exploratory data analysis report.'), S['body']))
        elems.append(Spacer(1, 0.4*cm))

        IMG_W = 7.4 * cm
        IMG_H = 5.0 * cm

        def plotly_to_image_bytes(fig_json, label):
            import matplotlib
            matplotlib.use('Agg')
            import matplotlib.pyplot as plt
            import numpy as np

            fig_mpl, ax = plt.subplots(figsize=(6, 3.6))
            ax.set_facecolor('#f8fafc')
            fig_mpl.patch.set_facecolor('white')

            try:
                traces = fig_json.get('data', [])
                layout = fig_json.get('layout', {})
                plotted = False

                for trace in traces:
                    ttype = trace.get('type', '')
                    name  = trace.get('name', '')
                    if ttype == 'histogram':
                        x = trace.get('x', [])
                        if x:
                            ax.hist(x, bins=20, color='#6366F1', alpha=0.75,
                                    edgecolor='white', linewidth=0.5)
                            plotted = True
                    elif ttype == 'box':
                        y = trace.get('y') or trace.get('x', [])
                        if y:
                            ax.boxplot(y, vert=bool(trace.get('y')), patch_artist=True,
                                       boxprops=dict(facecolor='#EEF2FF', color='#6366F1'),
                                       medianprops=dict(color='#1B2A4A', linewidth=2),
                                       whiskerprops=dict(color='#64748B'),
                                       capprops=dict(color='#64748B'),
                                       flierprops=dict(marker='o', color='#EF4444',
                                                       alpha=0.5, markersize=3))
                            plotted = True
                    elif ttype in ('scatter', 'scattergl'):
                        x = trace.get('x', [])
                        y = trace.get('y', [])
                        mode  = trace.get('mode', 'markers')
                        color = trace.get('marker', {}).get('color', '#6366F1')
                        if isinstance(color, list): color = '#6366F1'
                        if x and y:
                            if 'lines' in mode and 'markers' in mode:
                                ax.plot(x, y, color=color, alpha=0.8,
                                        linewidth=1.2, label=name)
                                ax.scatter(x, y, color=color, s=15, alpha=0.6, zorder=3)
                            elif 'lines' in mode:
                                ax.plot(x, y, color=color, alpha=0.8,
                                        linewidth=1.5, label=name)
                            else:
                                sz = trace.get('marker',{}).get('size', 5)
                                if isinstance(sz, list):
                                    sz = [max(2, min(s/4, 30)) for s in sz]
                                else:
                                    sz = max(3, min(sz, 20))
                                ax.scatter(x, y, c=color, s=sz,
                                           alpha=0.6, label=name)
                            plotted = True
                    elif ttype == 'bar':
                        x = trace.get('x', [])
                        y = trace.get('y', [])
                        color = trace.get('marker',{}).get('color','#6366F1')
                        if isinstance(color, list) and color:
                            color = color[0] if isinstance(color[0], str) else '#6366F1'
                        if x and y:
                            ax.bar(range(len(x)), y, color=color, alpha=0.8,
                                   edgecolor='white', linewidth=0.5, label=name)
                            ax.set_xticks(range(len(x)))
                            ax.set_xticklabels([str(xi)[:10] for xi in x],
                                               rotation=30, ha='right', fontsize=7)
                            plotted = True
                    elif ttype == 'pie':
                        vals = trace.get('values', [])
                        lbls = trace.get('labels', [])
                        if vals and lbls:
                            ax.pie(vals[:8], labels=[str(l)[:12] for l in lbls[:8]],
                                   autopct='%1.1f%%', colors=plt.cm.Set3.colors,
                                   textprops={'fontsize': 7})
                            plotted = True
                    elif ttype == 'heatmap':
                        z   = trace.get('z', [])
                        x   = trace.get('x', [])
                        y_a = trace.get('y', [])
                        if z:
                            z_arr = np.array(z, dtype=float)
                            im = ax.imshow(z_arr, cmap='RdYlGn', aspect='auto',
                                           vmin=-1, vmax=1)
                            if x:
                                ax.set_xticks(range(len(x)))
                                ax.set_xticklabels([str(xi)[:8] for xi in x],
                                                   rotation=45, ha='right', fontsize=6)
                            if y_a:
                                ax.set_yticks(range(len(y_a)))
                                ax.set_yticklabels([str(yi)[:8] for yi in y_a], fontsize=6)
                            plt.colorbar(im, ax=ax, shrink=0.8)
                            plotted = True
                    elif ttype == 'violin':
                        y = trace.get('y', [])
                        if y:
                            ax.violinplot([y], showmeans=True, showmedians=True)
                            plotted = True

                if not plotted:
                    ax.text(0.5, 0.5, f'{label}\n(Visualisasi tidak dapat dirender)',
                            ha='center', va='center', transform=ax.transAxes,
                            fontsize=10, color='#64748B',
                            bbox=dict(boxstyle='round,pad=0.5',
                                      facecolor='#f1f5f9', edgecolor='#cbd5e1'))
                    ax.set_xlim(0,1); ax.set_ylim(0,1)

                title_text = layout.get('title', {})
                if isinstance(title_text, dict):
                    title_text = title_text.get('text', label)
                ax.set_title(title_text or label, fontsize=9, fontweight='bold',
                             color='#1B2A4A', pad=6)
                xaxis = layout.get('xaxis', {})
                yaxis = layout.get('yaxis', {})
                if xaxis.get('title'):
                    xt = xaxis['title']
                    ax.set_xlabel(xt.get('text',xt) if isinstance(xt,dict) else xt,
                                  fontsize=7, color='#64748B')
                if yaxis.get('title'):
                    yt = yaxis['title']
                    ax.set_ylabel(yt.get('text',yt) if isinstance(yt,dict) else yt,
                                  fontsize=7, color='#64748B')
                ax.tick_params(labelsize=7)
                ax.spines['top'].set_visible(False)
                ax.spines['right'].set_visible(False)
                ax.spines['left'].set_color('#e2e8f0')
                ax.spines['bottom'].set_color('#e2e8f0')
                ax.grid(axis='y', color='#f1f5f9', linewidth=0.5, alpha=0.8)
                handles, lbs = ax.get_legend_handles_labels()
                if handles and len(handles) > 1:
                    ax.legend(fontsize=6, loc='upper right', framealpha=0.7)
            except Exception as inner_e:
                ax.text(0.5, 0.5, f'{label}\nError: {str(inner_e)[:50]}',
                        ha='center', va='center', transform=ax.transAxes,
                        fontsize=8, color='#EF4444')

            plt.tight_layout(pad=0.8)
            buf = io.BytesIO()
            fig_mpl.savefig(buf, format='png', dpi=130, bbox_inches='tight',
                            facecolor='white', edgecolor='none')
            plt.close(fig_mpl)
            buf.seek(0)
            return buf

        chart_items = list(chart_data.items())
        for idx in range(0, len(chart_items), 2):
            row_cells = []
            for key, fig_json in chart_items[idx:idx+2]:
                label = DL_CHART_LABELS.get(key, key)
                try:
                    if isinstance(fig_json, dict) and 'data' in fig_json:
                        buf = plotly_to_image_bytes(fig_json, label)
                        img = RLImage(buf, width=IMG_W, height=IMG_H)
                        row_cells.append([img, Paragraph(label, S['caption'])])
                    else:
                        row_cells.append([Paragraph(f'[{label}: ' + _T('data tidak tersedia','data not available') + ']',
                                                    S['caption'])])
                except Exception as e:
                    row_cells.append([Paragraph(f'[{label}: ' + _T('gagal render','render failed') + f' — {str(e)[:60]}]',
                                                S['caption'])])
            if len(row_cells) == 1:
                row_cells.append([Paragraph('', S['body'])])
            chart_row = Table(
                [[row_cells[0], row_cells[1]]],
                colWidths=[IMG_W + 0.3*cm, IMG_W + 0.3*cm],
            )
            chart_row.setStyle(TableStyle([
                ('VALIGN',       (0,0),(-1,-1), 'TOP'),
                ('LEFTPADDING',  (0,0),(-1,-1), 4),
                ('RIGHTPADDING', (0,0),(-1,-1), 4),
                ('TOPPADDING',   (0,0),(-1,-1), 6),
                ('BOTTOMPADDING',(0,0),(-1,-1), 10),
            ]))
            elems.append(chart_row)

    doc.build(elems)
    return filepath

# ============================================================
# EXPORT EXCEL
# ============================================================

def export_excel(df, col_meta, lang='id'):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    _en = lang == 'en'
    _T = lambda id_txt, en_txt: en_txt if _en else id_txt

    filepath = os.path.join(OUTPUT_DIR, 'exported_files', 'DataMind_Export.xlsx')
    wb = Workbook()

    def style_ws(ws, hdr_color):
        hf = Font(bold=True, color='FFFFFF', size=10, name='Calibri')
        fill = PatternFill('solid', fgColor=hdr_color)
        for cell in ws[1]:
            cell.font = hf
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 28)
        ws.row_dimensions[1].height = 20

    num_stats = get_numerical_stats(df, col_meta)
    cat_stats = get_categorical_stats(df, col_meta)

    ws1 = wb.active; ws1.title = _T('Data Preview', 'Data Preview')
    ws1.append(list(df.columns))
    for _, row in df.head(20).iterrows():
        ws1.append([str(v) if v is not None else '' for v in row])
    style_ws(ws1, '1B2A4A')

    ws2 = wb.create_sheet(_T('Statistik Numerik', 'Numerical Statistics'))
    ws2.append([_T('Variabel','Variable'), _T('Skala','Scale'), 'N', 'Mean', _T('Median','Median'),
                _T('Std Dev','Std Dev'), _T('Variance','Variance'), _T('Min','Min'), _T('Max','Max'),
                'Skewness', 'Kurtosis', _T('Outlier','Outlier'), _T('Missing','Missing'),
                _T('Missing%','Missing%'), _T('Normalitas','Normality'), 'CV(%)'])
    for r in num_stats:
        ws2.append([r.get(k,'') for k in ['column','scale','count','mean','median','std','variance','min','max','skewness','kurtosis','outliers','missing_count','missing_pct','normality','cv']])
    style_ws(ws2, '2563EB')

    ws3 = wb.create_sheet(_T('Statistik Kategorikal', 'Categorical Statistics'))
    ws3.append([_T('Variabel','Variable'), _T('Skala','Scale'), 'N', _T('Nilai Unik','Unique Values'),
                _T('Modus','Mode'), _T('Frek Modus','Mode Freq'), _T('Modus%','Mode%'),
                _T('Missing','Missing'), _T('Missing%','Missing%')])
    for r in cat_stats:
        ws3.append([r.get(k,'') for k in ['column','scale','count','unique','mode','mode_freq','mode_pct','missing_count','missing_pct']])
    style_ws(ws3, '059669')

    ws4 = wb.create_sheet(_T('Info Kolom', 'Column Info'))
    ws4.append([_T('Kolom','Column'), _T('Skala Pengukuran','Measurement Scale'), _T('Tipe Data','Data Type')])
    for col, scale in col_meta.items():
        ws4.append([col, scale, str(df[col].dtype) if col in df.columns else '-'])
    style_ws(ws4, '64748B')

    wb.save(filepath)
    return filepath


# ============================================================
# EXPORT HTML
# ============================================================

DL_CHART_LABELS = {
    'histogram':'Histogram','boxplot':'Boxplot','density':'Density Plot',
    'qq_plot':'QQ Plot','violin':'Violin Plot','bar_chart':'Bar Chart',
    'pie_chart':'Pie Chart','count_plot':'Count Plot','pareto_chart':'Pareto Chart',
    'scatter':'Scatter Plot','corr_heatmap':'Correlation Heatmap',
    'regression':'Regression Plot','bubble':'Bubble Chart',
    'boxplot_cat':'Boxplot per Kategori','violin_cat':'Violin per Kategori',
    'grouped_bar':'Grouped Bar','strip_plot':'Strip Plot',
    'ts_line':'Time Series Line','ts_ma':'Moving Average',
    'ts_rolling':'Rolling Mean','ts_trend':'Trend Line',
}

def export_html(df, col_meta, pages=None, chart_data=None, lang='id'):
    if pages is None:
        pages = ['cover', 'tim', 'daftar_isi', 'bab1', 'bab2', 'bab3', 'bab4', 'vis']
    if chart_data is None:
        chart_data = {}

    num_stats  = get_numerical_stats(df, col_meta)
    cat_stats  = get_categorical_stats(df, col_meta)
    preview    = df.head(5).where(df.head(5).notna(), other=None)
    cols_list  = df.columns.tolist()

    total_rows    = len(df)
    total_cols    = len(df.columns)
    num_cols_cnt  = sum(1 for v in col_meta.values() if v in ['interval','ratio'])
    cat_cols_cnt  = sum(1 for v in col_meta.values() if v in ['nominal','ordinal'])
    dt_cols_cnt   = sum(1 for v in col_meta.values() if v == 'datetime')
    missing_cells = int(df.isnull().sum().sum())
    missing_pct   = round(missing_cells / max(df.size, 1) * 100, 2)
    generated_at  = datetime.now().strftime('%d %B %Y, %H:%M WIB')
    year_now      = datetime.now().year

    not_normal   = [r for r in num_stats if str(r.get('normality','')).lower().replace('_',' ') == 'not normal']
    normal_ok    = [r for r in num_stats if r.get('normality') == 'Normal']
    outlier_cols = [r for r in num_stats if float(r.get('outliers') or 0) > 0]
    skewed_pos   = [r for r in num_stats if float(r.get('skewness') or 0) >  1]
    skewed_neg   = [r for r in num_stats if float(r.get('skewness') or 0) < -1]
    high_miss    = [r for r in num_stats + cat_stats if float(r.get('missing_pct') or 0) > 10]
    low_miss     = [r for r in num_stats + cat_stats if 0 < float(r.get('missing_pct') or 0) <= 10]
    dominant_cat = [r for r in cat_stats
                    if str(r.get('mode_pct') or '0') not in ('-','','nan')
                    and float(str(r.get('mode_pct') or 0)) > 30]

    def fmt(v):
        if v is None or v == '' or v == '-': return '—'
        try:
            n = float(v)
        except (ValueError, TypeError):
            return str(v)[:14]
        if abs(n) >= 1e9: return f'{n/1e9:.2f}B'
        if abs(n) >= 1e6: return f'{n/1e6:.2f}M'
        if abs(n) >= 1e4: return f'{n:,.0f}'
        return f'{n:.4g}'

    # Language flag
    _is_en = lang == 'en'

    # Temuan utama
    findings = []
    if num_stats:
        hm = max(num_stats, key=lambda r: float(r.get('mean') or 0))
        findings.append(f'Variabel <b>{hm["column"]}</b> memiliki nilai rata-rata tertinggi ({fmt(hm.get("mean"))}), variabel dengan magnitudo terbesar dalam dataset.')
    if outlier_cols:
        mo = max(outlier_cols, key=lambda r: float(r.get('outliers') or 0))
        pct_ = round(float(mo.get('outliers',0))/max(total_rows,1)*100,1)
        findings.append(f'Variabel <b>{mo["column"]}</b> memiliki outlier terbanyak ({mo.get("outliers","")} data point / {pct_}%), memerlukan investigasi mendalam.')
    if not_normal:
        findings.append(f'Sebanyak <b>{len(not_normal)} variabel</b> tidak berdistribusi normal: {", ".join([r["column"] for r in not_normal[:4]])}. Disarankan menggunakan korelasi Spearman.')
    if dominant_cat:
        findings.append(f'Variabel <b>{dominant_cat[0]["column"]}</b> menunjukkan dominasi kategori "{dominant_cat[0].get("mode","")}" sebesar {dominant_cat[0].get("mode_pct","")}%.')
    findings.append(f'Rasio missing values <b>{missing_pct}%</b>, tergolong <b>{"sangat baik (<5%)" if missing_pct < 5 else "perlu penanganan sebelum pemodelan"}</b>.')

    findings_li = ''.join([f'<li style="margin-bottom:9px">{f}</li>' for f in findings])

    # Tabel anggota tim HTML
    tim_rows_html = ''.join([
        f'<tr><td style="text-align:center">{i}</td><td><b>{m["name"]}</b></td>'
        f'<td>{m["nim"]}</td><td>{m["role"]}</td></tr>'
        for i, m in enumerate(TEAM_MEMBERS, 1)
    ])

    num_json  = json.dumps(num_stats,  ensure_ascii=False)
    cat_json  = json.dumps(cat_stats,  ensure_ascii=False)
    cols_json = json.dumps(cols_list,  ensure_ascii=False)
    meta_json = json.dumps(col_meta,   ensure_ascii=False)
    prev_json = json.dumps(preview.to_dict(orient='records'), ensure_ascii=False, default=str)

    # Custom chart JSON data for embedding
    chart_data_json = json.dumps(chart_data, ensure_ascii=False) if chart_data else '{}'

    # ── Build HTML sections conditionally ──
    sections_html = ''

    if 'cover' in pages:
        sections_html += f'''
<div class="cover" id="cover">
  <div class="cover-title">{'EXPLORATORY DATA ANALYSIS REPORT' if _is_en else 'LAPORAN ANALISIS DATA EKSPLORASI'}</div>
  <div class="cover-sub">Exploratory Data Analysis (EDA)</div>
  <table class="cover-table">
    <tr><th>{'Description' if _is_en else 'Keterangan'}</th><th>Detail</th></tr>
    <tr><td>{'Total Observations' if _is_en else 'Total Observasi'}</td><td><b>{total_rows:,} {'rows' if _is_en else 'baris'}</b></td></tr>
    <tr><td>{'Total Variables' if _is_en else 'Total Variabel'}</td><td><b>{total_cols} {'columns' if _is_en else 'kolom'}</b></td></tr>
    <tr><td>{'Numeric Variables' if _is_en else 'Variabel Numerik'}</td><td>{num_cols_cnt} {'columns' if _is_en else 'kolom'}</td></tr>
    <tr><td>{'Categorical Variables' if _is_en else 'Variabel Kategorikal'}</td><td>{cat_cols_cnt} {'columns' if _is_en else 'kolom'}</td></tr>
    <tr><td>{'Datetime Variables' if _is_en else 'Variabel Datetime'}</td><td>{dt_cols_cnt} {'columns' if _is_en else 'kolom'}</td></tr>
    <tr><td>Missing Values</td><td>{missing_cells:,} {'cells' if _is_en else 'sel'} ({missing_pct}%)</td></tr>
    <tr><td>{'Report Date' if _is_en else 'Tanggal Laporan'}</td><td>{generated_at}</td></tr>
  </table>
  <div class="cover-footer">
    {'Created with' if _is_en else 'Dibuat dengan'} <b>DataMind — Intelligent EDA Platform</b><br>
    {INSTITUSI} &mdash; {PRODI} &mdash; {KELOMPOK}<br>
    {'Lecturer' if _is_en else 'Dosen'}: {DOSEN} &mdash; &copy; {TAHUN} DataMind
  </div>
</div>
'''

    if 'tim' in pages:
        sections_html += f'''
<div class="chapter" id="tim">
  <div class="bab-label">''' + ('Page 1 — Development Team' if _is_en else 'Halaman 1 — Tim Pengembang') + '''</div>
  <div class="bab-title">''' + ('Development Team' if _is_en else 'Tim Pengembang') + '''</div>
  <hr class="bab-hr"/>
  <p style="margin-bottom:18px">''' + (f'DataMind was developed by students of {PRODI}, {INSTITUSI} as the final project for the Data Science Programming course in {TAHUN}.' if _is_en else f'DataMind dikembangkan oleh mahasiswa {PRODI}, {INSTITUSI} sebagai proyek akhir mata kuliah Data Science Programming tahun {TAHUN}.') + '''</p>

  <table class="dtbl navy" style="margin-bottom:6px">
    <thead><tr><th>''' + ('Information' if _is_en else 'Keterangan') + '''</th><th>Detail</th></tr></thead>
    <tbody>
      <tr><td>''' + ('Institution' if _is_en else 'Institusi') + '''</td><td>{INSTITUSI}</td></tr>
      <tr><td>''' + ('Study Program' if _is_en else 'Program Studi') + '''</td><td>{PRODI}</td></tr>
      <tr><td>''' + ('Course' if _is_en else 'Mata Kuliah') + '''</td><td>Data Science Programming</td></tr>
      <tr><td>''' + ('Group' if _is_en else 'Kelompok') + '''</td><td>{KELOMPOK}</td></tr>
      <tr><td>''' + ('Supervising Lecturer' if _is_en else 'Dosen Pembimbing') + '''</td><td>{DOSEN}</td></tr>
      <tr><td>''' + ('Academic Year' if _is_en else 'Tahun Akademik') + '''</td><td>{TAHUN}</td></tr>
    </tbody>
  </table>
  <div class="tbl-caption">''' + ('Table Tim.1 Institution and course information' if _is_en else 'Tabel Tim.1 Informasi institusi dan mata kuliah') + '''</div>

  <div class="sub">''' + ('Team Members' if _is_en else 'Anggota Tim') + '''</div>
  <table class="dtbl blue-team">
    <thead><tr><th style="width:40px">No.</th><th>''' + ('Full Name' if _is_en else 'Nama Lengkap') + '''</th><th>NIM</th><th>''' + ('Main Contribution' if _is_en else 'Kontribusi Utama') + '''</th></tr></thead>
    <tbody>{tim_rows_html}</tbody>
  </table>
  <div class="tbl-caption">''' + ('Table Tim.2 Team members and their respective contributions' if _is_en else 'Tabel Tim.2 Anggota tim dan kontribusi masing-masing') + '''</div>

  <div class="sub">''' + ('About DataMind' if _is_en else 'Tentang DataMind') + '''</div>
  <p>''' + ('DataMind is a web-based Exploratory Data Analysis platform that supports dataset upload (CSV, XLSX, TXT), automated statistical analysis, 20+ types of interactive visualizations, data quality detection and handling (missing, duplicate, outlier), time series analytics, AI-powered insight generation, and report export to PDF, Excel, and HTML formats. All features are designed to simplify the EDA process for both data scientists and general users.' if _is_en else 'DataMind adalah platform analisis data eksplorasi (<i>Exploratory Data Analysis</i>) berbasis web yang mendukung upload dataset (CSV, XLSX, TXT), analisis statistik otomatis, 20+ jenis visualisasi interaktif, deteksi dan penanganan data quality (missing, duplicate, outlier), time series analytics, AI-powered insight generation, serta ekspor laporan ke format PDF, Excel, dan HTML. Seluruh fitur dirancang untuk memudahkan proses EDA bagi data scientist maupun pengguna awam.') + '''</p>
</div>
'''

    if 'daftar_isi' in pages:
        toc_entries = '<tr class="toc-bab"><td>Cover</td><td class="toc-page">Cover</td></tr>'
        if 'tim' in pages:
            toc_entries += '<tr class="toc-bab"><td>' + ('Development Team' if _is_en else 'Tim Pengembang') + '</td><td class="toc-page">1</td></tr>'
        toc_entries += '<tr class="toc-bab"><td>' + ('Table of Contents' if _is_en else 'Daftar Isi') + '</td><td class="toc-page">2</td></tr>' + '''
    <tr class="toc-bab"><td>''' + ('CHAPTER I — Introduction and Dataset Overview' if _is_en else 'BAB I — Pendahuluan dan Gambaran Umum Dataset') + '''</td><td class="toc-page">3</td></tr>
    <tr class="toc-sub"><td>''' + ('1.1 Background' if _is_en else '1.1 Latar Belakang') + '''</td><td class="toc-page">3</td></tr>
    <tr class="toc-sub"><td>''' + ('1.2 Analysis Objectives' if _is_en else '1.2 Tujuan Analisis') + '''</td><td class="toc-page">3</td></tr>
    <tr class="toc-sub"><td>''' + ('1.3 Dataset Summary' if _is_en else '1.3 Ringkasan Dataset') + '''</td><td class="toc-page">3</td></tr>
    <tr class="toc-sub"><td>''' + ('1.4 Data Preview (First 5 Rows)' if _is_en else '1.4 Pratinjau Data (5 Baris Pertama)') + '''</td><td class="toc-page">3</td></tr>
    <tr class="toc-bab"><td>''' + ('CHAPTER II — Descriptive Statistics Numerical Variables' if _is_en else 'BAB II — Statistik Deskriptif Variabel Numerik') + '''</td><td class="toc-page">4</td></tr>
    <tr class="toc-sub"><td>''' + ('2.1 Numerical Statistics Summary' if _is_en else '2.1 Ringkasan Statistik Numerik') + '''</td><td class="toc-page">4</td></tr>
    <tr class="toc-sub"><td>''' + ('2.2 Distribution and Normality Analysis' if _is_en else '2.2 Analisis Distribusi dan Normalitas') + '''</td><td class="toc-page">4</td></tr>
    <tr class="toc-sub"><td>''' + ('2.3 Outlier Analysis' if _is_en else '2.3 Analisis Outlier') + '''</td><td class="toc-page">4</td></tr>
    <tr class="toc-bab"><td>''' + ('CHAPTER III — Categorical Variables & Data Quality' if _is_en else 'BAB III — Variabel Kategorikal & Kualitas Data') + '''</td><td class="toc-page">5</td></tr>
    <tr class="toc-sub"><td>''' + ('3.1 Categorical Variable Analysis' if _is_en else '3.1 Analisis Variabel Kategorikal') + '''</td><td class="toc-page">5</td></tr>
    <tr class="toc-sub"><td>''' + ('3.2 Category Dominance Patterns' if _is_en else '3.2 Pola Dominasi Kategori') + '''</td><td class="toc-page">5</td></tr>
    <tr class="toc-sub"><td>''' + ('3.3 Comprehensive Data Quality Analysis' if _is_en else '3.3 Analisis Kualitas Data Komprehensif') + '''</td><td class="toc-page">5</td></tr>
    <tr class="toc-bab"><td>''' + ('CHAPTER IV — Key Findings & Strategic Recommendations' if _is_en else 'BAB IV — Temuan Utama & Rekomendasi Strategis') + '''</td><td class="toc-page">6</td></tr>
    <tr class="toc-sub"><td>''' + ('4.1 Summary of Key Findings' if _is_en else '4.1 Ringkasan Temuan Utama') + '''</td><td class="toc-page">6</td></tr>
    <tr class="toc-sub"><td>''' + ('4.2 Recommendations for Further Analysis' if _is_en else '4.2 Rekomendasi Analisis Lanjutan') + '''</td><td class="toc-page">6</td></tr>'''
        if 'tim' in pages:
            toc_entries += '<tr class="toc-bab"><td>' + ('Development Team' if _is_en else 'Tim Pengembang') + '</td><td class="toc-page">7</td></tr>'
        if 'vis' in pages or chart_data:
            toc_entries += '<tr class="toc-bab"><td>' + ('Appendix — Selected Data Visualizations' if _is_en else 'Lampiran — Visualisasi Data Pilihan') + '</td><td class="toc-page">Lamp.</td></tr>'
        sections_html += f'''
<div class="chapter" id="daftar-isi">
  <div class="bab-label">''' + ('Page 2' if _is_en else 'Halaman 2') + '''</div>
  <div class="bab-title">''' + ('Table of Contents' if _is_en else 'Daftar Isi') + '''</div>
  <hr class="bab-hr"/>
  <table class="toc-table">
    {toc_entries}
  </table>
  <p style="margin-top:18px">''' + (f'This report was automatically generated by <b>DataMind EDA Platform</b> on {generated_at}.' if _is_en else f'Laporan ini dibuat otomatis oleh <b>DataMind EDA Platform</b> pada {generated_at}.') + '''</p>
</div>
'''

    if 'bab1' in pages:
        sections_html += f'''
<div class="chapter" id="bab1">
  <div class="bab-label">{'CHAPTER I' if _is_en else 'BAB I'}</div>
  <div class="bab-title">{'Introduction and Dataset Overview' if _is_en else 'Pendahuluan dan Gambaran Umum Dataset'}</div>
  <hr class="bab-hr"/>

  <div class="sub">''' + ('1.1 Background' if _is_en else '1.1 Latar Belakang') + '''</div>
  <p>''' + (f'This report is prepared as a result of exploratory data analysis (EDA) using the DataMind — Intelligent EDA Platform developed by {KELOMPOK}, {PRODI}, {INSTITUSI}. The analysis was conducted on a dataset with <b>{total_rows:,} observations</b> and <b>{total_cols} variables</b>, consisting of {num_cols_cnt} numerical variables, {cat_cols_cnt} categorical variables, and {dt_cols_cnt} datetime variables. EDA is a crucial stage in the data science cycle that aims to understand the structure, distribution, quality, and hidden patterns within the data before statistical modeling or machine learning is performed.' if _is_en else f'Laporan ini disusun sebagai hasil analisis eksplorasi data (<i>Exploratory Data Analysis</i>/EDA) menggunakan platform DataMind — Intelligent EDA Platform yang dikembangkan oleh {KELOMPOK}, {PRODI}, {INSTITUSI}. Analisis dilakukan terhadap dataset yang memiliki <b>{total_rows:,} observasi</b> dan <b>{total_cols} variabel</b>, terdiri dari {num_cols_cnt} variabel numerik, {cat_cols_cnt} variabel kategorikal, dan {dt_cols_cnt} variabel datetime. EDA merupakan tahapan krusial dalam siklus ilmu data yang bertujuan memahami struktur, distribusi, kualitas, dan pola tersembunyi di dalam data sebelum dilakukan pemodelan statistik maupun machine learning.') + '''</p>

  <div class="sub">''' + ('1.2 Analysis Objectives' if _is_en else '1.2 Tujuan Analisis') + '''</div>
  <p>''' + ('The main objectives of this report are: (1) to identify descriptive statistical characteristics of each variable; (2) to detect and document data quality issues such as missing values and outliers; (3) to analyze the distribution and normality of numerical data; (4) to understand patterns and frequencies in categorical variables; and (5) to provide strategic recommendations for further analysis that can be followed up by the research team or stakeholders.' if _is_en else 'Tujuan utama laporan ini adalah: (1) mengidentifikasi karakteristik statistik deskriptif setiap variabel; (2) mendeteksi dan mendokumentasikan masalah kualitas data seperti missing values dan outlier; (3) menganalisis distribusi dan normalitas data numerik; (4) memahami pola dan frekuensi pada variabel kategorikal; serta (5) memberikan rekomendasi strategis untuk analisis lanjutan yang dapat ditindaklanjuti oleh tim riset maupun pemangku kepentingan.') + '''</p>

  <div class="sub">''' + ('1.3 Dataset Summary' if _is_en else '1.3 Ringkasan Dataset') + '''</div>
  <div class="stat-cards">
    <div class="stat-card"><div class="stat-val">{total_rows:,}</div><div class="stat-lbl">''' + ('Total Rows' if _is_en else 'Total Baris') + '''</div></div>
    <div class="stat-card"><div class="stat-val">{total_cols}</div><div class="stat-lbl">''' + ('Total Columns' if _is_en else 'Total Kolom') + '''</div></div>
    <div class="stat-card"><div class="stat-val">{missing_pct}%</div><div class="stat-lbl">Missing Values</div></div>
    <div class="stat-card"><div class="stat-val">{num_cols_cnt}</div><div class="stat-lbl">''' + ('Num. Var.' if _is_en else 'Var. Numerik') + '''</div></div>
    <div class="stat-card"><div class="stat-val">{cat_cols_cnt}</div><div class="stat-lbl">''' + ('Cat. Var.' if _is_en else 'Var. Kategorikal') + '''</div></div>
    <div class="stat-card"><div class="stat-val">{dt_cols_cnt}</div><div class="stat-lbl">''' + ('Dt. Var.' if _is_en else 'Var. Datetime') + '''</div></div>
  </div>
  <table class="dtbl navy">
    <thead><tr><th>''' + ('Metric' if _is_en else 'Metrik') + '''</th><th>''' + ('Value' if _is_en else 'Nilai') + '''</th><th>''' + ('Description' if _is_en else 'Keterangan') + '''</th></tr></thead>
    <tbody>
      <tr><td>''' + ('Total Rows' if _is_en else 'Total Baris') + '''</td><td><b>{total_rows:,}</b></td><td>''' + ('Number of observations/records' if _is_en else 'Jumlah observasi/record') + '''</td></tr>
      <tr><td>''' + ('Total Columns' if _is_en else 'Total Kolom') + '''</td><td><b>{total_cols}</b></td><td>''' + ('Number of features/variables' if _is_en else 'Jumlah fitur/variabel') + '''</td></tr>
      <tr><td>''' + ('Numerical Variables' if _is_en else 'Variabel Numerik') + '''</td><td>{num_cols_cnt}</td><td>''' + ('Interval & ratio scale' if _is_en else 'Skala interval & rasio') + '''</td></tr>
      <tr><td>''' + ('Categorical Variables' if _is_en else 'Variabel Kategorikal') + '''</td><td>{cat_cols_cnt}</td><td>''' + ('Nominal & ordinal scale' if _is_en else 'Skala nominal & ordinal') + '''</td></tr>
      <tr><td>''' + ('Datetime Variables' if _is_en else 'Variabel Datetime') + '''</td><td>{dt_cols_cnt}</td><td>''' + ('Date/time type' if _is_en else 'Tipe tanggal/waktu') + '''</td></tr>
      <tr><td>''' + ('Total Missing' if _is_en else 'Total Missing') + '''</td><td>{missing_cells:,} ''' + ('cells' if _is_en else 'sel') + '''</td><td>{missing_pct}% ''' + ('of total data' if _is_en else 'dari total data') + '''</td></tr>
    </tbody>
  </table>
  <div class="tbl-caption">''' + ('Table 1.1 Summary of dataset dimensions and variable types' if _is_en else 'Tabel 1.1 Ringkasan dimensi dan tipe variabel dataset') + '''</div>

  <div class="sub">''' + ('1.4 Data Preview (First 5 Rows)' if _is_en else '1.4 Pratinjau Data (5 Baris Pertama)') + '''</div>
  <div class="tbl-scroll">
    <table class="dtbl" id="prev-tbl" style="font-size:8.5pt"></table>
  </div>
  <div class="tbl-caption">''' + ('Table 1.2 Preview of the first five rows of the dataset' if _is_en else 'Tabel 1.2 Pratinjau lima baris pertama dataset') + '''</div>
</div>
'''

    if 'bab2' in pages:
        sections_html += f'''
<div class="chapter" id="bab2">
  <div class="bab-label">{'CHAPTER II' if _is_en else 'BAB II'}</div>
  <div class="bab-title">{'Descriptive Statistics — Numerical Variables' if _is_en else 'Statistik Deskriptif Variabel Numerik'}</div>
  <hr class="bab-hr"/>

  <div class="sub">''' + ('2.1 Numerical Statistics Summary' if _is_en else '2.1 Ringkasan Statistik Numerik') + '''</div>
  <p>''' + (f'The following table presents complete descriptive statistics for <b>{num_cols_cnt} numerical variables</b>. Large values are formatted concisely (M=millions, B=billions) to avoid truncation. The table is responsive and can be scrolled horizontally on narrow screens.' if _is_en else f'Tabel berikut menyajikan statistik deskriptif lengkap untuk <b>{num_cols_cnt} variabel numerik</b>. Nilai besar diformat ringkas (M=juta, B=miliar) agar tidak terpotong. Tabel dibuat responsif sehingga dapat di-scroll secara horizontal di layar sempit.') + '''</p>
  <div class="tbl-scroll">
    <table class="dtbl" style="font-size:8.5pt;min-width:700px">
      <thead><tr>
        <th>''' + ('Variable' if _is_en else 'Variabel') + '''</th><th>''' + ('Scale' if _is_en else 'Skala') + '''</th><th>N</th><th>Mean</th><th>Median</th>
        <th>Std Dev</th><th>Min</th><th>Max</th><th>Skewness</th><th>Outlier</th><th>''' + ('Normality' if _is_en else 'Normalitas') + '''</th>
      </tr></thead>
      <tbody id="nb"></tbody>
    </table>
  </div>
  <div class="tbl-caption">''' + ('Table 2.1 Descriptive statistics of numerical variables (complete, untruncated)' if _is_en else 'Tabel 2.1 Statistik deskriptif variabel numerik (lengkap, tidak terpotong)') + '''</div>

  <div class="sub">''' + ('2.2 Distribution and Normality Analysis' if _is_en else '2.2 Analisis Distribusi dan Normalitas') + '''</div>
  <p id="dist-text"></p>

  <div class="sub">''' + ('2.3 Outlier Analysis' if _is_en else '2.3 Analisis Outlier') + '''</div>
  <p id="outlier-text"></p>
  <div id="outlier-tbl-wrap"></div>
</div>
'''

    if 'bab3' in pages:
        sections_html += f'''
<div class="chapter" id="bab3">
  <div class="bab-label">{'CHAPTER III' if _is_en else 'BAB III'}</div>
  <div class="bab-title">{'Descriptive Statistics — Categorical Variables & Data Quality' if _is_en else 'Statistik Deskriptif Variabel Kategorikal dan Kualitas Data'}</div>
  <hr class="bab-hr"/>

  <div class="sub">''' + ('3.1 Categorical Variable Analysis' if _is_en else '3.1 Analisis Variabel Kategorikal') + '''</div>
  <p>''' + (f'The dataset has <b>{cat_cols_cnt} categorical variables</b> analyzed in terms of frequency, category distribution, and mode dominance.' if _is_en else f'Dataset memiliki <b>{cat_cols_cnt} variabel kategorikal</b> yang dianalisis dari aspek frekuensi, distribusi kategori, dan dominasi nilai modus.') + '''</p>
  <div class="tbl-scroll">
    <table class="dtbl green" style="font-size:8.5pt">
      <thead><tr>
        <th>''' + ('Variable' if _is_en else 'Variabel') + '''</th><th>''' + ('Scale' if _is_en else 'Skala') + '''</th><th>N</th><th>''' + ('Unique' if _is_en else 'Nilai Unik') + '''</th>
        <th>''' + ('Mode' if _is_en else 'Modus') + '''</th><th>''' + ('Freq.' if _is_en else 'Frek.') + '''</th><th>''' + ('Mode%' if _is_en else 'Modus%') + '''</th><th>Missing%</th>
      </tr></thead>
      <tbody id="cb"></tbody>
    </table>
  </div>
  <div class="tbl-caption">''' + ('Table 3.1 Descriptive statistics of categorical variables' if _is_en else 'Tabel 3.1 Statistik deskriptif variabel kategorikal') + '''</div>

  <div class="sub">''' + ('3.2 Category Dominance Patterns' if _is_en else '3.2 Pola Dominasi Kategori') + '''</div>
  <p id="dom-text"></p>

  <div class="sub">''' + ('3.3 Comprehensive Data Quality Analysis' if _is_en else '3.3 Analisis Kualitas Data Komprehensif') + '''</div>
  <p id="quality-text"></p>
</div>
'''

    if 'bab4' in pages:
        sections_html += f'''
<div class="chapter" id="bab4">
  <div class="bab-label">{'CHAPTER IV' if _is_en else 'BAB IV'}</div>
  <div class="bab-title">{'Key Findings & Strategic Recommendations' if _is_en else 'Temuan Utama dan Rekomendasi Strategis'}</div>
  <hr class="bab-hr"/>

  <div class="sub">''' + ('4.1 Summary of Key Findings' if _is_en else '4.1 Ringkasan Temuan Utama') + '''</div>
  <p>''' + ('The following are key findings from the EDA process covering descriptive statistics, data quality, distribution, and categorical patterns.' if _is_en else 'Berikut temuan utama dari proses EDA yang mencakup statistik deskriptif, kualitas data, distribusi, dan pola kategorikal.') + '''</p>
  <div class="findings-box">
    <ul>{findings_li}</ul>
  </div>

  <div class="sub">''' + ('4.2 Recommendations for Further Analysis' if _is_en else '4.2 Rekomendasi Analisis Lanjutan') + '''</div>
  <div class="rec-item"><div class="rec-title">4.2.1 ''' + ('Data Preprocessing' if _is_en else 'Pra-pemrosesan Data') + '''</div>
    <p>''' + ('Handle outliers with winsorization or log transformation for skewed variables. Impute missing values: median for numerical, mode for categorical. Drop columns/rows if missing >30%.' if _is_en else 'Tangani outlier dengan winsorization atau transformasi log untuk variabel skewed. Imputasi missing values: median untuk numerik, modus untuk kategorikal. Drop kolom/baris jika missing >30%.') + '''</p></div>
  <div class="rec-item"><div class="rec-title">4.2.2 ''' + ('Correlation Analysis' if _is_en else 'Analisis Korelasi') + '''</div>
    <p>''' + ('Use Pearson correlation for normal variables and Spearman for non-normal ones. Visualize correlation heatmap to detect multicollinearity before modeling.' if _is_en else 'Gunakan korelasi Pearson untuk variabel normal dan Spearman untuk non-normal. Visualisasikan heatmap korelasi untuk mendeteksi multikolinearitas sebelum pemodelan.') + '''</p></div>
  <div class="rec-item"><div class="rec-title">4.2.3 ''' + ('Segmentation and Clustering' if _is_en else 'Segmentasi dan Pengelompokan') + '''</div>
    <p>''' + ('Leverage categorical variables for segmentation analysis (K-Means/clustering). Compare numerical distributions across segments to find meaningful patterns.' if _is_en else 'Manfaatkan variabel kategorikal untuk analisis segmentasi (K-Means/clustering). Bandingkan distribusi numerik antar segmen untuk menemukan pola bermakna.') + '''</p></div>
  <div class="rec-item"><div class="rec-title">4.2.4 ''' + ('Predictive Modeling' if _is_en else 'Pemodelan Prediktif') + '''</div>
    <p>''' + ('Apply normalization/standardization to numerical variables before modeling. For classification with class imbalance, use SMOTE or class_weight balancing.' if _is_en else 'Terapkan normalisasi/standarisasi pada variabel numerik sebelum pemodelan. Untuk klasifikasi dengan class imbalance, gunakan SMOTE atau class_weight balancing.') + '''</p></div>
  <div class="rec-item"><div class="rec-title">4.2.5 ''' + ('Validation and Reproducibility' if _is_en else 'Validasi dan Reproduktibilitas') + '''</div>
    <p>''' + ('Document all preprocessing decisions in a reproducible pipeline. Perform cross-validation (k-fold) to ensure good model generalization.' if _is_en else 'Dokumentasikan semua keputusan pra-pemrosesan dalam pipeline yang dapat direproduksi. Lakukan cross-validation (k-fold) untuk memastikan generalisasi model yang baik.') + '''</p></div>
</div>
'''

    # ═══ LAMPIRAN VISUALISASI ═══
    viz_html = ''
    if 'vis' in pages or chart_data:
        if chart_data:
            # Render custom charts from provided JSON data
            chart_items_json = json.dumps(list(chart_data.items()), ensure_ascii=False)
            viz_html = f'''
<div class="chapter viz-section" id="visualisasi">
  <div class="bab-label">''' + ('Appendix — Visualizations' if _is_en else 'Lampiran — Visualisasi') + '''</div>
  <div class="bab-title">''' + ('Selected Data Visualizations' if _is_en else 'Visualisasi Data Pilihan') + '''</div>
  <hr class="bab-hr"/>
  <p>''' + (f'This appendix contains <b>{len(chart_data)} visualizations</b> selected as supplementary material for the EDA report.' if _is_en else f'Lampiran ini memuat <b>{len(chart_data)} visualisasi</b> yang dipilih sebagai pelengkap laporan EDA.') + '''</p>
  <div class="viz-grid" id="vizGridCustom"></div>
</div>
<script>
(function() {{
  var chartItems = {chart_items_json};
  var grid = document.getElementById('vizGridCustom');
  if (!grid || !chartItems.length) return;
  chartItems.forEach(function(item) {{
    var key = item[0], fig = item[1];
    if (!fig || fig.error) return;
    var labels = {json.dumps(DL_CHART_LABELS, ensure_ascii=False)};
    var label = labels[key] || key;
    var plotId = 'plt-custom-' + key;
    var wrap = document.createElement('div');
    wrap.className = 'viz-item';
    wrap.innerHTML = '<div class="viz-title">' + label + '</div><div class="viz-plot" id="' + plotId + '"></div>';
    grid.appendChild(wrap);
    try {{
      Plotly.newPlot(plotId, fig.data, Object.assign({{}}, fig.layout, {{
        autosize: true, height: 240,
        showlegend: true
      }}), {{responsive: true, displayModeBar: false}});
    }} catch(e) {{
      document.getElementById(plotId).innerHTML = '<div style="display:flex;align-items:center;justify-content:center;height:100%;color:#94A3B8;font-size:10pt">''' + ('Visualization not available' if _is_en else 'Visualisasi tidak tersedia') + '''</div>';
    }}
  }});
}})();
</script>
'''
        else:
            # Fallback: generate synthetic charts from stats
            viz_html = '''
<div class="chapter viz-section" id="visualisasi">
  <div class="bab-label">''' + ('Appendix — Visualizations' if _is_en else 'Lampiran — Visualisasi') + '''</div>
  <div class="bab-title">''' + ('Selected Data Visualizations' if _is_en else 'Visualisasi Data Pilihan') + '''</div>
  <hr class="bab-hr"/>
  <p>''' + ('This appendix contains exploratory visualizations from the dataset. Charts are displayed interactively using Plotly.' if _is_en else 'Lampiran ini memuat visualisasi eksploratori dari dataset. Grafik ditampilkan secara interaktif menggunakan Plotly.') + '''</p>
  <div id="vizLoadStatus" style="text-align:center;padding:20px;color:#6366F1;font-weight:600">⏳ ''' + ('Loading visualizations...' if _is_en else 'Memuat visualisasi...') + '''</div>
  <div class="viz-grid" id="vizGrid" style="display:none"></div>
</div>

<script>
(function() {
  window.addEventListener('load', function() {
    var grid = document.getElementById('vizGrid');
    var status = document.getElementById('vizLoadStatus');
    if (!grid || typeof N === 'undefined') return;

    var charts = [];

    var numCols = N.slice(0, 4);
    numCols.forEach(function(r) {
      charts.push({
        label: 'Histogram \u2014 ' + r.column,
        data: [{
          type: 'histogram', name: r.column, x: [], nbinsx: 20,
          marker: {color: '#6366F1', opacity: 0.75}
        }],
        layout: {
          title: {text: r.column, font: {size: 11}},
          annotations: [{
            text: 'Mean: ' + (parseFloat(r.mean)||0).toFixed(2) + ' | Std: ' + (parseFloat(r.std)||0).toFixed(2),
            xref:'paper', yref:'paper', x:0.5, y:1.08, showarrow:false, font:{size:9, color:'#64748B'}
          }],
          paper_bgcolor:'#FAFAFA', plot_bgcolor:'#FAFAFA',
          font:{size:10}, margin:{t:60,r:10,b:40,l:50}
        },
        useStats: {mean: parseFloat(r.mean)||0, std: parseFloat(r.std)||1, min: parseFloat(r.min)||0, max: parseFloat(r.max)||1, col: r.column}
      });
    });

    var catCols = C.slice(0, 4);
    catCols.forEach(function(r) {
      charts.push({
        label: 'Bar Chart \u2014 ' + r.column,
        data: [{
          type: 'bar', x: [r.mode || '-'], y: [parseFloat(r.mode_freq)||0],
          text: [(r.mode_pct||'0')+'%'], textposition: 'auto',
          marker: {color: '#0D9488'}
        }],
        layout: {
          title: {text: r.column + ' (Top Kategori)', font: {size: 11}},
          paper_bgcolor:'#FAFAFA', plot_bgcolor:'#FAFAFA',
          font:{size:10}, margin:{t:40,r:10,b:60,l:50}
        }
      });
    });

    if (!charts.length) { status.textContent = 'No visualization data available.'; return; }
    status.style.display = 'none';
    grid.style.display = 'grid';

    charts.forEach(function(ch) {
      var item = document.createElement('div');
      item.className = 'viz-item';
      var plotId = 'plt-' + Math.random().toString(36).slice(2);
      item.innerHTML = '<div class="viz-title">' + ch.label + '</div><div class="viz-plot" id="' + plotId + '"></div>';
      grid.appendChild(item);
      if (ch.useStats) {
        var s = ch.useStats;
        var pts = [];
        for (var i = 0; i < 200; i++) {
          var u1 = Math.random(), u2 = Math.random();
          var z = Math.sqrt(-2 * Math.log(u1)) * Math.cos(2 * Math.PI * u2);
          var v = s.mean + z * s.std;
          if (v >= s.min && v <= s.max) pts.push(v);
        }
        ch.data[0].x = pts;
      }
      try {
        Plotly.newPlot(plotId, ch.data, Object.assign({}, ch.layout, {
          autosize: true, height: 240, showlegend: false
        }), {responsive: true, displayModeBar: false});
      } catch(e) {
        document.getElementById(plotId).innerHTML = '<div style="display:flex;align-items:center;justify-content:center;height:100%;color:#94A3B8;font-size:10pt">''' + ('Visualization not available' if _is_en else 'Visualisasi tidak tersedia') + '''</div>';
      }
    });
  });
})();
</script>
'''
        sections_html += viz_html

    # Footer
    sections_html += f'''
<div class="report-footer">
  {f'This report was automatically generated by <b>DataMind Intelligent EDA Platform</b> on {generated_at}.' if _is_en else f'Laporan ini dibuat secara otomatis oleh <b>DataMind Intelligent EDA Platform</b> pada {generated_at}.'}<br>
  {INSTITUSI} &mdash; {PRODI} &mdash; {KELOMPOK} &mdash; {'Lecturer' if _is_en else 'Dosen'}: {DOSEN}
</div>
'''

    # ═══ Full HTML assembly ═══
    _is_en = lang == 'en'

    # Pre-build bilingual JS strings
    _TXT_NORMAL = 'Normal'
    _TXT_NOT_NORMAL = 'Not Normal' if _is_en else 'Tidak Normal'
    _TXT_NORM_START = 'Normality test results show <b>' if _is_en else 'Hasil uji normalitas menunjukkan <b>'
    _TXT_NORM_MID = ' out of ' if _is_en else ' dari '
    _TXT_NORM_END1 = ' numerical variables</b> are normally distributed, while <b>' if _is_en else ' variabel numerik</b> berdistribusi normal, sementara <b>'
    _TXT_NORM_END2 = ' variables</b> are not normal. ' if _is_en else ' variabel</b> tidak normal. '
    _TXT_NONNORM = 'Non-normal variables: <b>' if _is_en else 'Variabel tidak normal: <b>'
    _TXT_NONNORM2 = 'Data transformation or non-parametric methods may be needed. ' if _is_en else 'Perlunya transformasi data atau metode non-parametrik. '
    _TXT_FOUND = 'Found ' if _is_en else 'Terdapat '
    _TXT_RSKEW = ' right-skewed variable' if _is_en else ' variabel right-skewed'
    _TXT_WHILE = 'While ' if _is_en else 'Sementara '
    _TXT_LSKEW = ' left-skewed variable' if _is_en else ' variabel left-skewed'
    _TXT_IQR = 'IQR analysis detected outliers in <b>' if _is_en else 'Analisis IQR mendeteksi outlier pada <b>'
    _TXT_IQR2 = ' numeric variables</b>. ' if _is_en else ' variabel numerik</b>. '
    _TXT_IQR_NONE = 'IQR analysis did not detect significant outliers. The dataset is clean from extreme values.' if _is_en else 'Analisis IQR tidak mendeteksi outlier yang signifikan. Dataset dalam kondisi bersih dari nilai ekstrem.'
    _TXT_VARIABLE = 'Variable' if _is_en else 'Variabel'
    _TXT_OUTLIER_CNT = 'Outlier Count' if _is_en else 'Jumlah Outlier'
    _TXT_PCT_TOTAL = '% of Total' if _is_en else '% dari Total'
    _TXT_RECOM = 'Recommendation' if _is_en else 'Rekomendasi'
    _TXT_FURTHER = 'Further investigation' if _is_en else 'Investigasi lebih lanjut'
    _TXT_TBL_OUT = 'Table 2.2 Summary of outliers per numeric variable' if _is_en else 'Tabel 2.2 Ringkasan outlier per variabel numerik'
    _TXT_DOM_FOUND = 'Found <b>' if _is_en else 'Terdapat <b>'
    _TXT_DOM_END = ' variables</b> with category dominance &gt;30%: ' if _is_en else ' variabel</b> dengan dominasi kategori &gt;30%: '
    _TXT_DOM_BY = 'dominated by' if _is_en else 'didominasi'
    _TXT_DOM_HIGH = 'High dominance may cause class imbalance in predictive modeling.' if _is_en else 'Dominasi tinggi berpotensi menyebabkan class imbalance dalam pemodelan prediktif.'
    _TXT_DOM_NONE = 'No extreme category dominance found (&gt;30%). Distribution among categories is relatively balanced.' if _is_en else 'Tidak ditemukan dominasi kategori ekstrem (&gt;30%). Distribusi antar kategori relatif seimbang.'
    _TXT_QUAL_START = 'Data quality evaluation was performed on <b>' if _is_en else 'Evaluasi kualitas data dilakukan terhadap <b>'
    _TXT_QUAL_MID = ' variables</b>. Total missing: <b>' if _is_en else ' variabel</b>. Total missing: <b>'
    _TXT_QUAL_CELLS = ' cells (' if _is_en else ' sel ('
    _TXT_QUAL_HIGH = 'Recommendation: median imputation (numeric) / mode (categorical), or drop if >30%.' if _is_en else 'Rekomendasi: imputasi median (numerik) / modus (kategorikal), atau drop jika >30%.'
    _TXT_QUAL_LOW = 'Low missing (1-10%): ' if _is_en else 'Missing rendah (1-10%): '
    _TXT_QUAL_LOW2 = 'Can be handled with simple imputation.' if _is_en else 'Dapat ditangani imputasi sederhana.'
    _TXT_QUAL_CLEAN = 'Dataset is clean, no missing values.' if _is_en else 'Dataset bersih, tidak ada missing values.'

    html = f"""<!DOCTYPE html>
<html lang="{lang}">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>DataMind — {'EDA Report' if _is_en else 'Laporan EDA'}</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Times New Roman',Times,serif;background:#f0f4f8;color:#1B2A4A;font-size:12pt;line-height:1.75}}
.page{{max-width:960px;margin:32px auto;padding:56px 72px;background:#fff;box-shadow:0 4px 40px rgba(0,0,0,.10);border-radius:4px}}

/* NAV / TOC sidebar */
.toc-nav{{position:fixed;top:24px;left:calc(50% + 480px + 12px);width:180px;background:#fff;
  border:1px solid #e2e8f0;border-radius:10px;padding:14px;font-size:9.5pt;
  box-shadow:0 2px 12px rgba(0,0,0,.08);display:none}}
@media(min-width:1340px){{.toc-nav{{display:block}}}}
.toc-nav a{{display:block;color:#64748B;text-decoration:none;padding:3px 0;border-bottom:1px solid #f1f5f9}}
.toc-nav a:hover{{color:#6366F1}}
.toc-nav-title{{font-weight:700;color:#1B2A4A;margin-bottom:8px;font-size:10pt}}

/* COVER */
.cover{{text-align:center;padding:52px 0 36px;border-bottom:4px solid #6366F1;margin-bottom:0}}
.cover-title{{font-size:22pt;font-weight:700;color:#6366F1;letter-spacing:.02em;margin-bottom:6px}}
.cover-sub{{font-size:13pt;color:#64748B;margin-bottom:28px}}
.cover-table{{margin:0 auto 28px;width:100%;max-width:520px;border-collapse:collapse;font-size:11pt}}
.cover-table th{{background:#1B2A4A;color:#fff;padding:9px 14px;text-align:left;font-weight:700}}
.cover-table td{{padding:8px 14px;border:1px solid #e2e8f0}}
.cover-table tr:nth-child(odd) td{{background:#F8FAFC}}
.cover-footer{{font-size:10pt;color:#94A3B8;margin-top:10px;line-height:1.6}}

/* CHAPTER */
.chapter{{margin-top:52px;padding-top:24px;border-top:1px solid #e2e8f0}}
.bab-label{{font-size:12pt;font-weight:700;color:#6366F1;text-transform:uppercase;letter-spacing:.06em;margin-bottom:2px}}
.bab-title{{font-size:14pt;font-weight:700;color:#1B2A4A;text-transform:uppercase;margin-bottom:10px}}
.bab-hr{{border:none;border-top:2.5px solid #6366F1;margin-bottom:18px}}
.sub{{font-size:12pt;font-weight:700;color:#1B2A4A;margin:20px 0 8px}}
p{{margin-bottom:12px;text-align:justify}}

/* STAT CARDS */
.stat-cards{{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin:18px 0}}
.stat-card{{background:#F8FAFC;border:1px solid #EEF2FF;border-radius:10px;padding:16px;text-align:center}}
.stat-val{{font-size:22pt;font-weight:800;color:#6366F1}}
.stat-lbl{{font-size:9pt;color:#64748B;text-transform:uppercase;letter-spacing:.05em;margin-top:4px}}

/* TABLES */
table.dtbl{{width:100%;border-collapse:collapse;font-size:9.5pt;margin:12px 0 6px;table-layout:fixed;word-wrap:break-word}}
table.dtbl th{{background:#6366F1;color:#fff;font-weight:700;padding:8px 8px;text-align:left;font-size:9pt}}
table.dtbl td{{padding:7px 8px;border-bottom:1px solid #e2e8f0;word-break:break-word}}
table.dtbl tr:nth-child(even) td{{background:#EEF2FF}}
table.dtbl tr:hover td{{background:#EEF2FF}}
table.dtbl.green th{{background:#059669}}
table.dtbl.green tr:nth-child(even) td{{background:#ECFDF5}}
table.dtbl.navy th{{background:#1B2A4A}}
table.dtbl.navy tr:nth-child(even) td{{background:#F8FAFC}}
table.dtbl.amber th{{background:#D97706}}
table.dtbl.amber tr:nth-child(even) td{{background:#FFFBEB}}
table.dtbl.blue-team th{{background:#6366F1}}
table.dtbl.blue-team tr:nth-child(even) td{{background:#EEF2FF}}
.tbl-caption{{font-size:9pt;color:#64748B;text-align:center;margin-bottom:18px;font-style:italic}}
.tbl-scroll{{overflow-x:auto}}

/* BADGE */
.badge{{display:inline-block;padding:2px 7px;border-radius:4px;font-size:8pt;font-weight:700}}
.badge-ratio{{background:rgba(16,185,129,.15);color:#059669}}
.badge-interval{{background:rgba(99,102,241,.15);color:#6366F1}}
.badge-nominal{{background:rgba(244,63,94,.15);color:#dc2626}}
.badge-ordinal{{background:rgba(245,158,11,.15);color:#d97706}}
.badge-datetime{{background:rgba(139,92,246,.15);color:#7c3aed}}
.normal-ok{{color:#059669;font-weight:700}}
.normal-no{{color:#dc2626;font-weight:700}}

/* FINDINGS */
.findings-box{{background:#EEF2FF;border-left:4px solid #6366F1;border-radius:0 10px 10px 0;padding:20px 24px;margin:10px 0 20px}}
.findings-box ul{{padding-left:18px}}

/* RECS */
.rec-item{{background:#F8FAFC;border:1px solid #e2e8f0;border-radius:10px;padding:16px 18px;margin-bottom:12px}}
.rec-title{{font-weight:700;color:#6366F1;margin-bottom:6px;font-size:11pt}}

/* TOC page */
.toc-table{{width:100%;border-collapse:collapse;font-size:11pt}}
.toc-table tr.toc-bab td{{font-weight:700;color:#1B2A4A;padding:8px 0 4px;border-bottom:1.5px solid #6366F1}}
.toc-table tr.toc-sub td{{color:#64748B;padding:3px 0 3px 18px;border-bottom:1px solid #f1f5f9;font-size:10pt}}
.toc-table td.toc-page{{text-align:right;white-space:nowrap;font-weight:700;color:#6366F1}}

/* TIM */
.tim-section{{background:#F8FAFC;border-radius:12px;padding:24px;margin-top:16px}}

/* VISUALISASI */
.viz-section{{margin-top:52px;padding-top:24px;border-top:1px solid #e2e8f0}}
.viz-grid{{display:grid;grid-template-columns:repeat(2,1fr);gap:20px;margin-top:16px}}
.viz-item{{border:1px solid #E2E8F0;border-radius:10px;padding:14px;background:#FAFAFA}}
.viz-title{{font-size:10pt;font-weight:700;color:#1B2A4A;margin-bottom:8px;text-align:center}}
.viz-plot{{height:260px;width:100%}}
@media print{{.viz-grid{{grid-template-columns:1fr}}}}

/* FOOTER */
.report-footer{{margin-top:40px;padding-top:16px;border-top:2.5px solid #6366F1;text-align:center;font-size:9.5pt;color:#94A3B8}}
@media print{{body{{background:white}}.page{{box-shadow:none;padding:24px 48px;margin:0}}.toc-nav{{display:none}}}}
</style>
<script src="https://cdn.plot.ly/plotly-2.35.2.min.js"></script>
</head>
<body>

<nav class="toc-nav">
  <div class="toc-nav-title">{'Navigation' if _is_en else 'Navigasi'}</div>
  <a href="#cover">Cover</a>
  <a href="#tim">{'Development Team' if _is_en else 'Tim Pengembang'}</a>
  <a href="#daftar-isi">{'Table of Contents' if _is_en else 'Daftar Isi'}</a>
  <a href="#bab1">{'CHAPTER I' if _is_en else 'BAB I'}</a>
  <a href="#bab2">{'CHAPTER II' if _is_en else 'BAB II'}</a>
  <a href="#bab3">{'CHAPTER III' if _is_en else 'BAB III'}</a>
  <a href="#bab4">{'CHAPTER IV' if _is_en else 'BAB IV'}</a>
  <a href="#visualisasi">{'Visualizations' if _is_en else 'Visualisasi'}</a>
</nav>

<div class="page">

{sections_html}

</div><!-- /page -->

<script>
var N={num_json}, C={cat_json}, M={meta_json}, K={cols_json}, P={prev_json};

function fmt(v) {{
  if (v===null||v===undefined||v===''||v==='-') return '—';
  var n = parseFloat(v);
  if (isNaN(n)) return String(v).substring(0,14);
  if (Math.abs(n)>=1e9) return (n/1e9).toFixed(2)+'B';
  if (Math.abs(n)>=1e6) return (n/1e6).toFixed(2)+'M';
  if (Math.abs(n)>=1e4) return n.toLocaleString('id-ID',{{maximumFractionDigits:0}});
  return parseFloat(n.toPrecision(4)).toString();
}}

// Pratinjau tabel — semua kolom, scroll horizontal
var pt = document.getElementById('prev-tbl');
var hdr = '<thead><tr>'+K.map(function(c){{
  return '<th style="white-space:nowrap">'+c+'<br><span class="badge badge-'+(M[c]||'nominal')+'">'+(M[c]||'nominal').toUpperCase()+'</span></th>';
}}).join('')+'</tr></thead>';
var bod = '<tbody>';
P.forEach(function(r){{
  bod += '<tr>'+K.map(function(c){{
    var v = r[c]; return '<td style="white-space:nowrap">'+(v!==null&&v!==undefined?v:'—')+'</td>';
  }}).join('')+'</tr>';
}});
bod += '</tbody>';
pt.innerHTML = hdr + bod;

// Numerical table
var nb = document.getElementById('nb');
N.forEach(function(r){{
  var isNorm = r.normality === 'Normal';
  nb.innerHTML += '<tr>'+
    '<td><b>'+r.column+'</b></td>'+
    '<td><span class="badge badge-'+r.scale+'">'+r.scale.toUpperCase()+'</span></td>'+
    '<td>'+(r.count||'—')+'</td>'+
    '<td>'+fmt(r.mean)+'</td><td>'+fmt(r.median)+'</td>'+
    '<td>'+fmt(r.std)+'</td><td>'+fmt(r.min)+'</td><td>'+fmt(r.max)+'</td>'+
    '<td>'+fmt(r.skewness)+'</td>'+
    '<td>'+(r.outliers||'0')+'</td>'+
    '<td class="'+(isNorm?'normal-ok':'normal-no')+'">'+(isNorm?'{_TXT_NORMAL}':'{_TXT_NOT_NORMAL}')+'</td>'+
  '</tr>';
}});

// Categorical table
var cb = document.getElementById('cb');
C.forEach(function(r){{
  cb.innerHTML += '<tr>'+
    '<td><b>'+r.column+'</b></td>'+
    '<td><span class="badge badge-'+r.scale+'">'+r.scale.toUpperCase()+'</span></td>'+
    '<td>'+(r.count||'—')+'</td><td>'+(r.unique||'—')+'</td>'+
    '<td>'+(r.mode||'—')+'</td><td>'+(r.mode_freq||'—')+'</td>'+
    '<td>'+(r.mode_pct||'—')+'%</td><td>'+(r.missing_pct||'0')+'%</td>'+
  '</tr>';
}});

// Distribution & normality text
var notN = N.filter(function(r){{return r.normality&&r.normality.toLowerCase().replace('_',' ').includes('not normal');}});
var normOk = N.filter(function(r){{return r.normality==='Normal';}});
var skPos  = N.filter(function(r){{return parseFloat(r.skewness||0)>1;}});
var skNeg  = N.filter(function(r){{return parseFloat(r.skewness||0)<-1;}});
var dt = '{_TXT_NORM_START}'+normOk.length+'{_TXT_NORM_MID}'+N.length+'{_TXT_NORM_END1}'+notN.length+'{_TXT_NORM_END2}';
if (notN.length) dt += '{_TXT_NONNORM}'+notN.slice(0,5).map(function(r){{return r.column;}}).join(', ')+'</b>. {_TXT_NONNORM2}';
if (skPos.length) dt += '{_TXT_FOUND}'+skPos.length+'{_TXT_RSKEW}'+(skPos.length>1?'s':'')+': '+skPos.slice(0,3).map(function(r){{return r.column;}}).join(', ')+'. ';
if (skNeg.length) dt += '{_TXT_WHILE}'+skNeg.length+'{_TXT_LSKEW}'+(skNeg.length>1?'s':'')+': '+skNeg.slice(0,3).map(function(r){{return r.column;}}).join(', ')+'.';
document.getElementById('dist-text').innerHTML = dt;

// Outlier text + table
var outCols = N.filter(function(r){{return parseFloat(r.outliers||0)>0;}});
var ot = '';
if (outCols.length) {{
  ot = '{_TXT_IQR}'+outCols.length+'{_TXT_IQR2}';
  var outHtml = '<div class="tbl-scroll"><table class="dtbl amber" style="font-size:9pt;margin-top:10px"><thead><tr><th>{_TXT_VARIABLE}</th><th>{_TXT_OUTLIER_CNT}</th><th>{_TXT_PCT_TOTAL}</th><th>{_TXT_RECOM}</th></tr></thead><tbody>';
  outCols.forEach(function(r){{
    var cnt  = parseInt(r.outliers||0);
    var pct_ = (cnt / {total_rows} * 100).toFixed(2);
    var rec  = pct_ > 5 ? 'Winsorize / Log Transform' : '{_TXT_FURTHER}';
    outHtml += '<tr><td><b>'+r.column+'</b></td><td>'+cnt+'</td><td>'+pct_+'%</td><td>'+rec+'</td></tr>';
  }});
  outHtml += '</tbody></table></div><div class="tbl-caption">{_TXT_TBL_OUT}</div>';
  document.getElementById('outlier-tbl-wrap').innerHTML = outHtml;
}} else {{
  ot = '{_TXT_IQR_NONE}';
}}
document.getElementById('outlier-text').innerHTML = ot;

// Category dominance
var domCat = C.filter(function(r){{return parseFloat(r.mode_pct||0)>30;}});
var domt = '';
if (domCat.length) {{
  domt = '{_TXT_DOM_FOUND}'+domCat.length+'{_TXT_DOM_END}';
  domCat.slice(0,4).forEach(function(r){{domt += '<b>'+r.column+'</b> {_TXT_DOM_BY} &ldquo;'+r.mode+'&rdquo; ('+r.mode_pct+'%); ';}});
  domt += '{_TXT_DOM_HIGH}';
}} else {{
  domt = '{_TXT_DOM_NONE}';
}}
document.getElementById('dom-text').innerHTML = domt;

// Data quality
var highM = [...N,...C].filter(function(r){{return parseFloat(r.missing_pct||0)>10;}});
var lowM  = [...N,...C].filter(function(r){{var p=parseFloat(r.missing_pct||0);return p>0&&p<=10;}});
var qt = '{_TXT_QUAL_START}{total_cols}{_TXT_QUAL_MID}{missing_cells:,}{_TXT_QUAL_CELLS}{missing_pct}%)</b>. ';
if (highM.length) qt += 'Missing &gt;10%: <b>'+highM.slice(0,5).map(function(r){{return r.column;}}).join(', ')+'</b>. {_TXT_QUAL_HIGH} ';
if (lowM.length)  qt += '{_TXT_QUAL_LOW}'+lowM.slice(0,5).map(function(r){{return r.column;}}).join(', ')+'. {_TXT_QUAL_LOW2} ';
if (!highM.length && !lowM.length) qt += '{_TXT_QUAL_CLEAN} ';
document.getElementById('quality-text').innerHTML = qt;
</script>
</body>
</html>"""

    filepath = os.path.join(OUTPUT_DIR, 'reports', 'DataMind_Dashboard.html')
    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(html)
    return filepath
